#!/usr/bin/env python3
# CAHPS Auditor
# Copyright (C) 2026 HST Pathways. All rights reserved. Developed by Tyler Brock. 
# This copyright notice must be preserved in all copies, modifications, and derivative works of this software.
import openpyxl
import os
import re
import sys
import urllib.parse
import uuid
import warnings
import webbrowser
from multiprocessing import Pool, cpu_count, freeze_support
from tqdm import tqdm
from audit_printer import save_report, build_report
from audit_lib_funcs import *
from audit_oas_funcs import (
    _CPT_LOAD_ERROR,
    classify_cpt,
    cpt_is_ineligible,
    calc_e_m_total,
    find_frame_inel_count,
    check_req_headers,
    validate_inel_repeat_rows,
)
from audit_hcahps_funcs import _DRG_APR_LOAD_ERROR

__version__ = "2.1.5"
version = __version__


def print_app_info_and_help_block():
    print(f"CAHPS Auditor v{version} by Tyler Brock")
    print()
    print(
        "Need help? Visit https://github.com/ToonLunk/OAS-CAHPS-Auditor, or contact support."
    )
    print()
    install_dir = os.path.join(os.environ.get("LOCALAPPDATA", "C:\\Users\\<user>\\AppData\\Local"), "OAS-CAHPS-Auditor")
    print("If there's a new version of this software, you will see an update notice here.")
    print("If you need to update SIDs or CPT code lists, you can find them below:")
    print(f"  - SIDs: {install_dir}\\SIDs.csv")
    print(f"  - CPT codes: {install_dir}\\cpt_codes.json")
    print(f"  - DRG/APR codes: {install_dir}\\drg_apr_codes.json")


def check_for_updates():
    """Check GitHub for latest version and notify if update available.
    
    Returns a dict with 'latest_version' and 'download_url' if an update is
    available, or None otherwise.
    """
    try:
        import urllib.request
        import json
        from packaging import version as pkg_version
        
        url = "https://api.github.com/repos/ToonLunk/OAS-CAHPS-Auditor/releases"
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'OAS-CAHPS-Auditor')
        
        with urllib.request.urlopen(req, timeout=3) as response:
            releases = json.loads(response.read().decode())
            if releases:
                latest_version = releases[0].get('tag_name', '').lstrip('v')
                
                # Only notify if the latest version is greater than current version
                if latest_version and pkg_version.parse(latest_version) > pkg_version.parse(version):
                    print(f"\nUpdate available: v{latest_version} (current: v{version})")
                    print(f"Download: https://github.com/ToonLunk/OAS-CAHPS-Auditor/releases/latest\n")
                    return {
                        'latest_version': latest_version,
                        'download_url': 'https://github.com/ToonLunk/OAS-CAHPS-Auditor/releases/latest',
                    }
    except Exception:
        # Silently fail if unable to check
        pass
    return None


def audit_excel(file_path, show_progress=False):
    try:
        if show_progress:
            print(f"Loading workbook: {os.path.basename(file_path)}...")
        # Suppress openpyxl's "serial value outside date limits" warnings — those cells
        # are already returned as None by openpyxl, and our blank-date checks flag them.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r".*serial value.*outside the limits for dates.*",
                category=UserWarning,
            )
            wb = openpyxl.load_workbook(file_path, data_only=True)
    except:
        print(
            f"--- Critical Error opening {file_path}! Are you sure it's an Excel file?"
        )
        input("Press enter to continue: ")
        print("\n")
        sys.exit(1)

    # Auto-detect audit type: HCAHPS workbooks contain a "CMS" tab; OAS do not.
    if "CMS" in wb.sheetnames:
        audit_type = "HCAHPS"
    else:
        audit_type = "OAS"

    # HCAHPS-specific processing and validation 
    if audit_type == "HCAHPS":
        from audit_hcahps_funcs import check_req_headers as hcahps_check_req_headers
        sheet = wb["CMS"]

        # --- Extract and clean header/footer ---
        raw_header = pick_header(sheet)
        raw_footer = pick_footer(sheet)

        header = clean_hf_text(raw_header)
        footer = clean_hf_text(raw_footer)

        # --- Extract values from header/footer text ---
        # HCAHPS does not include a SUBMITTED value in the header.
        patients_submitted = None

        header_clean = re.sub(r"&\[[^\]]+\]", "", header)
        _sig_matches = re.findall(r"(?<![A-Z])([A-Z]{2,3})(?![A-Z\d])", header_clean)
        two_letter_code = _sig_matches[-1] if _sig_matches else ""

        # HCAHPS SIDs have 2-letter prefixes (e.g. "AC123" not "ANM123")
        sid_match = re.search(r"([A-Z]{2}\d+)", header_clean)
        header_sid = sid_match.group(1) if sid_match else None

        if header_sid:
            prefix_match = re.match(r"([A-Z]{2})", header_sid)
            sid_prefix = prefix_match.group(1) if prefix_match else None
        else:
            sid_prefix = None

        # Look up client name from HCAHPS SID registry
        sid_registry_name = None
        if sid_prefix:
            sid_registry_name = lookup_sid_client_name(
                sid_prefix,
                sid_filename='HCAHPS_SIDs.csv',
                onedrive_link=HCAHPS_SIDS_ONEDRIVE_LINK,
                show_missing_warning=True,
                sid_col_idx=1,
            )

        basefname = os.path.basename(file_path)
        base_before_hash = basefname.split("#", 1)[0]

        nums = "".join(str(ord(c) - 64) for c in two_letter_code)
        uuid_code = uuid.uuid4().hex
        audit_id = f"{uuid_code}{nums}"

        el_match = re.search(r"EL\s*=\s*(\d+)", footer)
        ss_match = re.search(r"SS\s*=\s*(\d+)", footer)
        eligible_patients = int(el_match.group(1)) if el_match else None
        sample_size = int(ss_match.group(1)) if ss_match else None

        # --- Find column indexes ---
        first_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {cell.value: idx for idx, cell in enumerate(first_row, start=1)}

        mapping, missing_req_headers = hcahps_check_req_headers(headers, header_text=header)

        sid_col     = mapping.get("SID")
        mrn_col     = mapping.get("MRN")
        tel_col     = mapping.get("TELEPHONE")
        ddate_col   = mapping.get("D.DATE")
        age_col     = mapping.get("AGE")
        ds_col      = mapping.get("DS")
        gender_col  = mapping.get("GENDER")
        unit_col    = mapping.get("UNIT")
        physician_col = mapping.get("PHYSICIAN NAME")
        drg_col     = mapping.get("DRG")
        att_col     = mapping.get("ATT")
        lag_col     = mapping.get("LAG")
        id_col      = mapping.get("ID")
        fd_col      = mapping.get("FD")
        lg_col      = mapping.get("LG")
        email_col   = mapping.get("EMAIL ADDRESS")
        cms_col     = mapping.get("CMS INDICATOR")
        lang_col    = mapping.get("LANGUAGE")

        issues = []

        # Validate SID sequence (only for rows with CMS=1)
        sid_issues = []
        sid_row_issues = []
        if sid_col and cms_col:
            sid_issues, sid_row_issues = validate_sid_sequence(sheet, sid_col, cms_col, header_sid)  # type: ignore
            issues.extend(sid_issues)
            if show_progress:
                print(f"[OK] SID validation complete ({len(sid_issues)} issues found)")

        # Check for same-day discharges in FRAME tab (admit date == discharge date -> should be INEL)
        inel_issues = []
        inel_row_issues = None
        if 'FRAME' in wb.sheetnames:
            from audit_hcahps_funcs import check_same_day_discharges
            inel_issues, inel_row_issues = check_same_day_discharges(wb)
            issues.extend(inel_issues)
            if show_progress:
                found = len(inel_row_issues) if inel_row_issues else 0
                print(f"[OK] Same-day discharge check complete ({found} issues found)")

        # Detect FRAME columns and validate addresses
        frame_col_map = None
        frame_invalid_addresses = []
        frame_noted_addresses = []
        if 'FRAME' in wb.sheetnames:
            from audit_hcahps_funcs import get_frame_col_map, check_frame_addresses
            frame_col_map = get_frame_col_map(wb)
            frame_invalid_addresses, frame_noted_addresses = check_frame_addresses(wb, frame_col_map)
            if show_progress:
                print(f"[OK] FRAME address check complete ({len(frame_invalid_addresses)} issues found)")

        # Validate EXCLU tab
        exclu_count = None
        exclu_row_issues = []
        if "EXCLU" in wb.sheetnames:
            from audit_hcahps_funcs import validate_exclu_rows
            exclu_count, exclu_row_issues = validate_exclu_rows(wb["EXCLU"])
            if show_progress:
                print(f"[OK] EXCLU validation complete ({len(exclu_row_issues)} issues found)")

        # Validate INEL tab
        inel_count = None
        inel_tab_row_issues = []
        if "INEL" in wb.sheetnames:
            from audit_hcahps_funcs import validate_inel_rows
            inel_count, inel_tab_row_issues = validate_inel_rows(wb["INEL"])
            if show_progress:
                print(f"[OK] INEL validation complete ({len(inel_tab_row_issues)} issues found)")

        # Count DUP tab rows with 'D' in the DUP column
        dup_count = None
        if "DUP" in wb.sheetnames:
            from audit_hcahps_funcs import count_dup_d_rows
            dup_count = count_dup_d_rows(wb["DUP"])
            if show_progress:
                print(f"[OK] DUP tab count complete ({dup_count} 'D' rows found)")

        # Extract discharge date range and validate blank dates
        service_date_range = None
        blank_date_row_issues = []
        if ddate_col:
            service_date_range, blank_date_issues, blank_date_row_issues = extract_service_date_range(
                sheet, ddate_col, mrn_col=mrn_col, cms_col=cms_col
            )
            issues.extend(blank_date_issues)
            if show_progress:
                print(f"[OK] Discharge date extraction complete")

        # If filename specifies a day-range (e.g. "APRIL 16-30"), validate CMS discharge dates
        if ddate_col and '#' in basefname:
            _fn_name_part = os.path.splitext(basefname.split('#', 1)[1])[0]
            _fn_year = None
            for _t in _fn_name_part.split():
                try:
                    _y = int(_t)
                    if 2000 <= _y <= 2100:
                        _fn_year = _y
                        break
                except ValueError:
                    pass
            from audit_hcahps_funcs import parse_filename_date_range, check_cms_discharge_date_range
            _fn_start, _fn_end = parse_filename_date_range(
                _fn_name_part,
                filename_year=_fn_year,
                service_date_range=service_date_range,
            )
            if _fn_start is not None and _fn_end is not None:
                _, _date_range_row_issues = check_cms_discharge_date_range(
                    sheet, ddate_col, mrn_col, cms_col, _fn_start, _fn_end,
                )
                blank_date_row_issues.extend(_date_range_row_issues)
                if show_progress:
                    print(f"[OK] Filename date range check complete ({len(_date_range_row_issues)} issues found)")

        # Calculate name match status for batch reporting
        name_match_info = None
        if sid_prefix and sid_registry_name:
            normalized_registry = re.sub(r'\s*-?\s*\d{1,2}/\d{1,2}(?:/\d{2,4})?\s*$', '', sid_registry_name).strip().lower()
            normalized_filename = base_before_hash.strip().lower()
            names_match = (normalized_registry == normalized_filename)
            name_match_info = {
                'filename': base_before_hash,
                'registry_name': sid_registry_name,
                'match': names_match,
            }

        # Count CMS=1 and CMS=2 rows (HCAHPS has no E/M column)
        em_col = None
        total_em = None
        emails = None
        mailings = None
        cms1_count = 0
        non_reported = 0
        if cms_col:
            for _row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(cell is not None for cell in _row):
                    continue
                cms_val = _row[cms_col - 1]
                if cms_val is not None:
                    try:
                        cms_int = int(float(str(cms_val).strip()))
                        if cms_int == 1:
                            cms1_count += 1
                        elif cms_int == 2:
                            non_reported += 1
                    except (ValueError, TypeError):
                        pass

        # Look up facility/location names from POP tab
        facility_matches = []
        from audit_lib_funcs import FACILITY_NAME_ALIASES, find_all_columns_in_sheet
        if 'POP' in wb.sheetnames:
            tab_matches = find_all_columns_in_sheet(wb['POP'], FACILITY_NAME_ALIASES)
            for m in tab_matches:
                m['tab'] = 'POP'
            facility_matches.extend(tab_matches)

        if show_progress:
            print("Building report...")

        report_lines, issues = build_report(
            wb=wb,
            sheet=sheet,
            file_path=file_path,
            version=version,
            audit_id=audit_id,
            missing_req_headers=missing_req_headers,
            patients_submitted=None,
            eligible_patients=eligible_patients,
            sample_size=sample_size,
            sid_prefix=sid_prefix,
            sid_registry_name=sid_registry_name,
            emails=emails,
            mailings=mailings,
            total_em=total_em,
            non_reported=non_reported,
            cms1_count=cms1_count,
            headers=headers,
            issues=issues,
            count_nonempty_rows=count_nonempty_rows,
            classify_cpt=None,
            cpt_is_ineligible=None,
            addr1_col=None,
            addr2_col=None,
            city_col=None,
            state_col=None,
            zip_col=None,
            cms_col=cms_col,
            em_col=em_col,
            find_frame_inel_count=None,
            mrn_col=mrn_col,
            sid_col=sid_col,
            sid_row_issues=sid_row_issues,
            inel_row_issues=inel_row_issues,
            service_date_range=service_date_range,
            blank_date_row_issues=blank_date_row_issues,
            facility_matches=facility_matches,
            exclu_count=exclu_count,
            exclu_row_issues=exclu_row_issues,
            inel_count=inel_count,
            inel_tab_row_issues=inel_tab_row_issues,
            dup_count=dup_count,
            frame_col_map=frame_col_map,
            frame_invalid_addresses=frame_invalid_addresses,
            frame_noted_addresses=frame_noted_addresses,
            audit_type="HCAHPS",
            header_text=header,
        )

        if show_progress:
            print("[OK] Report built successfully")

        return file_path, report_lines, service_date_range, name_match_info

    # OAS-specific processing and validation
    sheet = wb["OASCAPHS"]

    # --- Extract and clean header/footer ---
    raw_header = pick_header(sheet)
    raw_footer = pick_footer(sheet)

    header = clean_hf_text(raw_header)
    footer = clean_hf_text(raw_footer)

    # --- Extract values from header/footer text ---
    submitted_match = re.search(r"SUBMITTED\s*=\s*(\d+)", header)
    patients_submitted = int(submitted_match.group(1)) if submitted_match else None

    # finds the two-letter code (like "TB") - can be anywhere in header
    header_clean = re.sub(r"&\[[^\]]+\]", "", header)
    # Match exactly 2 uppercase letters not part of a longer word
    m = re.search(r"(?<![A-Z])([A-Z]{2})(?![A-Z])", header_clean)
    two_letter_code = m.group(1) if m else ""

    # Extract SID from header (should be first SID in sequence)
    sid_match = re.search(r"([A-Z]{2,3}\d+)", header_clean)
    header_sid = sid_match.group(1) if sid_match else None
    
    # Extract SID prefix (2-3 letter code) for display
    if header_sid:
        prefix_match = re.match(r"([A-Z]{2,3})", header_sid)
        sid_prefix = prefix_match.group(1) if prefix_match else None
    else:
        sid_prefix = None
    
    # Look up client name from SID registry
    sid_registry_name = None
    if sid_prefix:
        sid_registry_name = lookup_sid_client_name(
            sid_prefix,
            sid_filename='SIDs.csv',
            onedrive_link=OAS_SIDS_ONEDRIVE_LINK,
            show_missing_warning=True,
        )
    
    # Look up facility/location names from FRAME tab (if it exists)
    # Look up facility/location names from POP tab only
    facility_matches = []
    from audit_lib_funcs import FACILITY_NAME_ALIASES, find_all_columns_in_sheet
    if 'POP' in wb.sheetnames:
        tab_matches = find_all_columns_in_sheet(wb['POP'], FACILITY_NAME_ALIASES)
        for m in tab_matches:
            m['tab'] = 'POP'
        facility_matches.extend(tab_matches)

    # Get base filename for comparison
    basefname = os.path.basename(file_path)
    base_before_hash = basefname.split("#", 1)[0]

    # convert letters to alphabet positions (A=1, B=2) and append to UUID
    nums = "".join(str(ord(c) - 64) for c in two_letter_code)
    uuid_code = uuid.uuid4().hex
    audit_id = f"{uuid_code}{nums}"

    el_match = re.search(r"EL\s*=\s*(\d+)", footer)
    ss_match = re.search(r"SS\s*=\s*(\d+)", footer)
    eligible_patients = int(el_match.group(1)) if el_match else None
    sample_size = int(ss_match.group(1)) if ss_match else None

    # --- Find column indexes ---
    first_row = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = {cell.value: idx for idx, cell in enumerate(first_row, start=1)}

    # Check for required headers (returns mapping and list of any missing)
    mapping, missing_req_headers = check_req_headers(headers)

    sid_col = mapping.get("SID")
    pat_col = mapping.get("PATIENT NAME")
    addr1_col = mapping.get("ADDRESS1")
    addr2_col = mapping.get("ADDRESS2")
    city_col = mapping.get("CITY")
    state_col = mapping.get("STATE")
    zip_col = mapping.get("ZIP")
    tel_col = mapping.get("TELEPHONE")
    svc_col = mapping.get("SERVICE DATE")
    gender_col = mapping.get("GENDER")
    age_col = mapping.get("AGE")
    mrn_col = mapping.get("MRN")
    surg_cat_col = mapping.get("SURGICAL CATEGORY")
    att_col = mapping.get("ATT")
    lag_col = mapping.get("LAG")
    id_col = mapping.get("ID")
    fd_col = mapping.get("FD")
    lg_col = mapping.get("LG")
    em_col = mapping.get("E/M")
    email_col = mapping.get("EMAIL ADDRESS")
    cms_col = mapping.get("CMS INDICATOR")
    lang_col = mapping.get("SURVEY LANGUAGE")

    issues = []

    # Validate SID sequence (only for rows with CMS=1) if columns exist
    sid_issues = []
    sid_row_issues = []
    if sid_col and cms_col:
        sid_issues, sid_row_issues = validate_sid_sequence(sheet, sid_col, cms_col, header_sid)  # type: ignore
        issues.extend(sid_issues)
        if show_progress:
            print(f"[OK] SID validation complete ({len(sid_issues)} issues found)")

    # Validate INEL tab REPEAT entries
    inel_issues = []
    inel_row_issues = []
    if "INEL" in wb.sheetnames:
        inel_sheet = wb["INEL"]
        inel_issues, inel_row_issues = validate_inel_repeat_rows(inel_sheet, show_progress=show_progress)
        issues.extend(inel_issues)
        if show_progress:
            print(f"[OK] INEL validation complete ({len(inel_issues)} issues found)")

    # Extract service date range and validate blank dates
    service_date_range = None
    blank_date_row_issues = []
    if svc_col:
        service_date_range, blank_date_issues, blank_date_row_issues = extract_service_date_range(
            sheet, svc_col, mrn_col=mrn_col, cms_col=cms_col
        )
        issues.extend(blank_date_issues)
        if show_progress:
            print(f"[OK] Service date extraction complete")

    # Calculate E/M totals if columns exist
    total_em = None
    emails = None
    mailings = None
    non_reported = None
    cms1_count = None
    
    if cms_col and em_col:
        try:
            total_em, emails, mailings, non_reported, cms1_count = calc_e_m_total(
                sheet, cms_col, em_col
            )  # type: ignore
        except Exception as e:
            issues.append(f"Error calculating E/M totals: {str(e)}")

    if show_progress:
        print("Building report...")
    report_lines, issues = build_report(
        wb=wb,
        sheet=sheet,
        file_path=file_path,
        version=version,
        audit_id=audit_id,
        missing_req_headers=missing_req_headers,
        patients_submitted=patients_submitted,
        eligible_patients=eligible_patients,
        sample_size=sample_size,
        sid_prefix=sid_prefix,
        sid_registry_name=sid_registry_name,
        emails=emails,
        mailings=mailings,
        total_em=total_em,
        non_reported=non_reported,
        cms1_count=cms1_count,
        headers=headers,
        issues=issues,
        count_nonempty_rows=count_nonempty_rows,
        classify_cpt=classify_cpt,
        cpt_is_ineligible=cpt_is_ineligible,
        addr1_col=addr1_col,
        addr2_col=addr2_col,
        city_col=city_col,
        state_col=state_col,
        zip_col=zip_col,
        cms_col=cms_col,
        em_col=em_col,
        find_frame_inel_count=find_frame_inel_count,  # optional
        mrn_col=mrn_col,  # optional
        sid_col=sid_col,  # SID column
        sid_row_issues=sid_row_issues,  # SID validation issues
        inel_row_issues=inel_row_issues,  # INEL validation issues
        service_date_range=service_date_range,  # Service date range
        blank_date_row_issues=blank_date_row_issues,  # Blank date issues
        facility_matches=facility_matches,  # Facility/location columns from FRAME and POP tabs
    )
    
    if show_progress:
        print("[OK] Report built successfully")
    
    # Calculate name match status for batch reporting
    name_match_info = None
    if sid_prefix and sid_registry_name:
        # Normalize both names for comparison (same logic as in audit_printer)
        # Make case-insensitive and remove date pattern (with optional dash) from end
        normalized_registry = re.sub(r'\s*-?\s*\d{1,2}/\d{1,2}(?:/\d{2,4})?\s*$', '', sid_registry_name).strip().lower()
        normalized_filename = base_before_hash.strip().lower()
        names_match = (normalized_registry == normalized_filename)
        name_match_info = {
            'filename': base_before_hash,
            'registry_name': sid_registry_name,
            'match': names_match
        }

    return file_path, report_lines, service_date_range, name_match_info


def process_file_wrapper(args):
    """Wrapper function for multiprocessing to process a single Excel file.
    
    Args:
        args: Tuple of (filename, version_str, update_info)
        
    Returns:
        dict with status, filename, result_file, name_match_info, and error (if any)
    """
    filename, version_str, update_info = args
    try:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            file_path, report_lines, service_date_range, name_match_info = audit_excel(filename)
            final_file = save_report(file_path, report_lines, version=version_str, service_date_range=service_date_range, update_info=update_info)
        warning_msgs = [str(w.message) for w in caught]
        return {
            'status': 'success',
            'filename': filename,
            'result_file': final_file,
            'name_match_info': name_match_info,
            'warnings': warning_msgs,
            'error': None
        }
    except Exception as e:
        return {
            'status': 'error',
            'filename': filename,
            'result_file': None,
            'name_match_info': None,
            'warnings': [],
            'error': str(e)
        }


# Module-level update info (set in __main__ before any auditing)
_update_info = None

if __name__ == "__main__":
    # Required for PyInstaller multiprocessing support on Windows
    freeze_support()
    
    _update_info = check_for_updates()

    if _CPT_LOAD_ERROR:
        print("=" * 60)
        print("  WARNING: cpt_codes.json could not be loaded.")
        print(f"  {_CPT_LOAD_ERROR}")
        print("  CPT ineligibility checks will use built-in defaults.")
        print("=" * 60)
        print()

    if _DRG_APR_LOAD_ERROR:
        print("=" * 60)
        print("  WARNING: drg_apr_codes.json could not be loaded.")
        print(f"  {_DRG_APR_LOAD_ERROR}")
        print("  DRG/APR ineligibility checks will use built-in defaults.")
        print("=" * 60)
        print()

    remaining_argv = sys.argv[1:]

    if len(remaining_argv) != 1:
        print("Usage: audit <excel_file> or audit --all")
        print("Options:")
        print("  --all       Process all Excel files in the current directory")
        print("  --help,-h   Show this help message")
        print("  --version,-v Show version information")
        print("\n")
        print(
            "Need help? Visit https://github.com/ToonLunk/OAS-CAHPS-Auditor, or contact support."
        )
        sys.exit(1)

    arg = remaining_argv[0]

    # Handle --all flag to process all files in current directory
    if arg == "--all":
        print_app_info_and_help_block()
        print()
        # Get list of Excel files
        excel_files = [
            f for f in os.listdir(".") if f.endswith((".xlsx", ".xls", ".xlsm"))
        ]

        if not excel_files:
            print("No Excel files found in current directory.")
            sys.exit(0)

        files_processed = 0
        name_mismatch_files = []  # Track files with name mismatches
        num_processes = min(cpu_count(), len(excel_files))
        print(f"Found {len(excel_files)} Excel file(s) to process.")
        print(f"Using {num_processes} processor(s) for parallel processing.\n")

        # Prepare arguments for worker function (filename, version, update_info)
        worker_args = [(f, version, _update_info) for f in excel_files]

        # Process files in parallel with progress bar
        # Using imap_unordered with chunksize=1 for immediate feedback
        with Pool(processes=num_processes) as pool:
            results = list(tqdm(
                pool.imap_unordered(process_file_wrapper, worker_args, chunksize=1),
                total=len(excel_files),
                desc="Processing files",
                unit="file",
                smoothing=0  # Disable smoothing for immediate updates
            ))
        
        # Process results
        for result in results:
            if result['status'] == 'success':
                print(f"[OK] {result['filename']} -> {result['result_file']}")
                files_processed += 1
                
                # Track name mismatch if applicable
                name_match_info = result['name_match_info']
                if name_match_info and not name_match_info['match']:
                    name_mismatch_files.append({
                        'filename': result['filename'],
                        'file_name': name_match_info['filename'],
                        'registry_name': name_match_info['registry_name']
                    })
            else:
                print(f"[ERROR] {result['filename']}: {result['error']}")

        print(
            f"\nCompleted: {files_processed}/{len(excel_files)} file(s) processed successfully."
        )
        
        # Print name matching summary
        print("\n" + "="*60)
        print("CLIENT NAME MATCHING SUMMARY")
        print("="*60)
        if name_mismatch_files:
            print(f"\n !!!  {len(name_mismatch_files)} file(s) with CLIENT NAME MISMATCH:\n")
            for item in name_mismatch_files:
                print(f"  File: {item['filename']}")
                print(f"    - Filename:      {item['file_name']}")
                print(f"    - Registry Name: {item['registry_name']}")
                print()
        else:
            print("\n[OK] All files have matching client names (or no registry data)\n")
        print("="*60 + "\n")

        # Collect any processing warnings from workers
        all_warnings = []
        for result in results:
            for msg in result.get('warnings', []):
                all_warnings.append(f"  [{result['filename']}] {msg}")
        if all_warnings:
            print("WARNINGS DURING PROCESSING:")
            for w in all_warnings:
                print(w)
            print()
        input("Press Enter to exit: ")
        sys.exit(0)

    if arg == "--help" or arg == "-h":
        print("Usage: audit <excel_file> or audit --all [--lookup]")
        print("Options:")
        print("  --all       Process all Excel files in the current directory")
        print("  --lookup    Append a people-search section for invalid emails / missing phones")
        print("  --help,-h   Show this help message")
        print("  --version,-v Show version information")
        print("\n")
        print_app_info_and_help_block()

        sys.exit(0)
    if arg == "--version" or arg == "-v":
        print(f"CAHPS Auditor version {version}")
        sys.exit(0)

    # Handle single file
    file_path = arg
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)

    try:
        print_app_info_and_help_block()
        print()
        print(f"Processing: {os.path.basename(file_path)}")
        with warnings.catch_warnings(record=True) as _caught_warnings:
            warnings.simplefilter("always")
            file_path, report_lines, service_date_range, name_match_info = audit_excel(file_path, show_progress=False)
            final_file = save_report(file_path, report_lines, version=version, service_date_range=service_date_range, update_info=_update_info)
        print(f"Report saved: {final_file}")

        # Open the report in the default browser.
        # For long network paths Windows routes webbrowser.open() to Edge
        # instead of the default browser.  Skip auto-open in that case and
        # show an obvious message instead — report is still saved normally.
        _report_url = 'file:///' + final_file.replace('\\', '/').replace('#', '%23')
        _long_path = len(final_file) >= 200
        if _long_path:
            print()
            print("=" * 60)
            print("  NOTE: Report was NOT auto-opened.")
            print("  The file path is too long for reliable auto-open.")
            print("  Navigate to the AUDITS folder and open it manually:")
            print(f"  {final_file}")
            print("=" * 60)
        else:
            try:
                webbrowser.open(_report_url)
                print("Opening report in your default browser...")
            except Exception as e:
                print(f"Could not automatically open browser: {e}")

        # Print clickable link for easy access
        print(f"\nReport link: {_report_url}")

        if _caught_warnings:
            print()
            print("WARNINGS DURING PROCESSING:")
            for _w in _caught_warnings:
                print(f"  {_w.message}")
            print()
            input("Press Enter to exit: ")
        elif _long_path:
            input("Press Enter to exit: ")

    except Exception as e:
        # For single file mode, print error and exit
        print(f"\nError processing file: {e}")
        import traceback
        traceback.print_exc()
        input("Press Enter to exit: ")
        sys.exit(1)
