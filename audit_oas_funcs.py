# OAS CAHPS Auditor - OAS-specific functions
# Copyright (C) 2026 HST Pathways. All rights reserved.
# Originally developed by Tyler Brock. This copyright notice, including
# authorship credit to Tyler Brock, must be preserved in all copies,
# modifications, and derivative works of this software.
#
# This module contains functions that are specific to OAS CAHPS auditing:
#   - CPT code loading and validation
#   - E/M and CMS INDICATOR totalling
#   - FRAME tab ineligibility counting
#   - OAS required header list
#
# Shared validation utilities (phone, email, address, dates, etc.) live in
# audit_lib_funcs.py and are imported by this module where needed.
import json
import os
import sys
from openpyxl.worksheet.worksheet import Worksheet


# --- CPT ineligibility rules (loaded from JSON) ---

def _get_cpt_config_path():
    """Get the path to cpt_codes.json from the installation directory."""
    if getattr(sys, 'frozen', False):
        # Running as PyInstaller bundle - use installation directory
        exe_dir = os.path.dirname(sys.executable)
        return os.path.join(exe_dir, 'cpt_codes.json')
    else:
        # Running as script
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, 'cpt_codes.json')


_CPT_LOAD_ERROR = None  # Set if cpt_codes.json fails to load; checked at startup in audit.py


def _load_cpt_config():
    """Load CPT code configuration from JSON file."""
    global _CPT_LOAD_ERROR
    config_path = _get_cpt_config_path()
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        return {
            'valid_ranges': config.get('valid_ranges', []),
            'invalid_ranges': config.get('invalid_ranges', []),
            'valid_codes': set(str(c).upper() for c in config.get('valid_codes', [])),
            'invalid_codes': set(str(c).upper() for c in config.get('invalid_codes', []))
        }
    except Exception as e:
        _CPT_LOAD_ERROR = f"{config_path}: {e}"
        # Return defaults if file not found
        return {
            'valid_ranges': [[10004, 69990], [93451, 93462], [93566, 93572], [93985, 93986]],
            'invalid_ranges': [],
            'valid_codes': set(),
            'invalid_codes': set()
        }


_CPT_CONFIG = _load_cpt_config()
EXPLICIT_VALID_SET = _CPT_CONFIG['valid_codes']
INVALID_CPT_SET = _CPT_CONFIG['invalid_codes']


def cpt_is_ineligible(cpt_raw) -> tuple[bool, str]:
    """
    Determine whether a CPT code is ineligible.
    Returns (is_ineligible, reason).
    """
    if cpt_raw is None:
        return False, "blank CPT is OK"
    cpt = str(cpt_raw).strip().upper()
    if cpt == "":
        return False, "blank CPT is OK"

    # Exact explicit valid codes are always valid
    if cpt in EXPLICIT_VALID_SET:
        return False, "explicitly valid"

    # Exact invalid list
    if cpt in INVALID_CPT_SET:
        return True, "explicitly invalid list"

    # If purely numeric, check valid ranges
    if cpt.isdigit():
        num = int(cpt)

        # Check invalid ranges first
        for range_pair in _CPT_CONFIG['invalid_ranges']:
            if range_pair[0] <= num <= range_pair[1]:
                return True, "numeric in invalid range"

        # Check valid ranges
        for range_pair in _CPT_CONFIG['valid_ranges']:
            if range_pair[0] <= num <= range_pair[1]:
                return False, "numeric in valid range"

        return True, "outside valid ranges"

    # Non-numeric codes that are not in EXPLICIT_VALID_SET are ineligible
    return True, "not explicitly valid, and not in ranges"


def classify_cpt(cpt_code: str) -> int:
    """Return expected surgical category for a CPT code."""
    if not cpt_code:
        return 5
    txt = str(cpt_code).strip().lower()
    # Exact text codes
    if txt in ("g0105", "g0121", "g0104"):
        return 1
    if txt == "g0260":
        return 2
    # Numeric ranges
    if txt.isdigit():
        num = int(txt)
        if 40490 <= num <= 49999:
            return 1
        elif 20000 <= num <= 29999:
            return 2
        elif 65091 <= num <= 68999:
            return 3
        elif (
            (10004 <= num <= 19999)
            or (30000 <= num <= 39999)
            or (50000 <= num <= 64999)
            or (68900 <= num <= 69990)
            or (92920 <= num <= 93986)
        ):
            return 4
    return 5


def calc_e_m_total(sheet, cms_col, em_col):
    """Tally Email, Mailing, and non-reported counts from the OASCAPHS tab."""
    emails = 0
    mailings = 0
    non_reported = 0
    cms1_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        cms_val = row[cms_col - 1]  # type: ignore
        em_val = row[em_col - 1]  # type: ignore

        # normalize cms_val to an int when possible
        try:
            cms_num = int(cms_val) if cms_val is not None else None
        except (ValueError, TypeError):
            cms_num = None

        # normalize em_val to uppercase string for reliable comparison
        em_str = str(em_val).strip().upper() if em_val is not None else ""

        if cms_num == 1:
            cms1_count += 1
            if em_str == "E":
                emails += 1
            elif em_str == "M":
                mailings += 1
        elif cms_num == 2:
            non_reported += 1

    total_em = emails + mailings
    return total_em, emails, mailings, non_reported, cms1_count


def find_frame_inel_count(
    frame_sheet: Worksheet,
    top_nonempty_threshold: int = 3,
    min_block_rows: int = 3,
    max_blank_within_block: int = 1,
) -> int:
    """
    Locate the lower sparse block in the FRAME tab and count non-empty values
    in column B for that block (6-month repeat ineligibles).
    Returns integer count.
    """
    rows = list(frame_sheet.iter_rows(values_only=True))
    if not rows:
        return 0

    # count non-empty cells per row
    nonempty_counts = [
        sum(1 for c in row if c is not None and str(c).strip() != "") for row in rows
    ]

    # find last dense row
    last_dense_index = -1
    for i, cnt in enumerate(nonempty_counts):
        if cnt >= top_nonempty_threshold:
            last_dense_index = i

    # candidate start of sparse region
    start_idx = last_dense_index + 1
    if start_idx >= len(rows):
        return 0

    # accumulate a sparse run starting at start_idx allowing a small number of blanks inside
    sparse_run = 0
    blanks_in_run = 0
    i = start_idx
    while i < len(rows):
        cnt = nonempty_counts[i]
        if cnt <= 2:
            sparse_run += 1
            i += 1
        else:
            if sparse_run >= min_block_rows:
                break
            sparse_run = 0
            start_idx = i + 1
            i += 1

    if sparse_run < min_block_rows:
        # fallback: scan whole sheet for any run of rows with <=2 non-empty values
        start_idx = None
        for i in range(len(rows)):
            if nonempty_counts[i] <= 2 and nonempty_counts[i] != 0:
                run = 1
                blanks = 0
                for j in range(i + 1, len(rows)):
                    if nonempty_counts[j] == 0:
                        blanks += 1
                        if blanks > max_blank_within_block:
                            break
                    elif nonempty_counts[j] <= 2:
                        run += 1
                    else:
                        break
                if run >= min_block_rows:
                    start_idx = i
                    sparse_run = run
                    break
        if start_idx is None:
            return 0

    end_idx = start_idx + sparse_run  # exclusive

    # Count rows with any non-empty value from column B onwards in the sparse block.
    # Column A is skipped because RATSTATS random numbers may be placed there and
    # should not be counted as INEL entries. Identifiers may be in any column >= B.
    pt_id_count = 0
    for r in range(start_idx, end_idx):
        row = rows[r]
        if any(val is not None and str(val).strip() != "" for val in row[1:]):
            pt_id_count += 1

    return pt_id_count


def check_req_headers(headers):
    """
    Check that all required OAS CAHPS headers are present.
    Returns (mapping, missing_req_headers).
    """
    required_names = [
        "SID",
        "PATIENT NAME",
        "ADDRESS1",
        "CITY",
        "STATE",
        "ZIP",
        "TELEPHONE",
        "SERVICE DATE",
        "GENDER",
        "AGE",
        "PROVIDER NAME",
        "MRN",
        "P.TYPE",
        "SURGICAL CATEGORY",
        "ATT",
        "LAG",
        "ID",
        "FD",
        "LG",
        "E/M",
        "EMAIL ADDRESS",
        "CMS INDICATOR",
        "SURVEY LANGUAGE",
    ]

    mapping = {}
    missing_req_headers = []

    for name in required_names:
        mapping[name] = headers.get(name)
        if mapping[name] is None:
            missing_req_headers.append(name)

    # Return mapping and list of missing headers without raising exception
    return mapping, missing_req_headers


def validate_inel_repeat_rows(inel_sheet, show_progress=False):
    """
    Validate OAS INEL tab REPEAT entries.

    For rows marked as REPEAT (duplicates):
    - All cells in the row should have red font (RGB 255, 0, 0)
    - "REPEAT" should appear in the rightmost column
    - The "REPEAT" cell should have yellow background fill and bold red font
    - No other cells should have highlighting (yellow background)

    For rows with no cell-level highlighting:
    - They MUST have "REPEAT" marker, otherwise there's no indication why they're in INEL

    Returns (issues, row_issues) lists.
    """
    issues = []
    row_issues = []

    if inel_sheet is None:
        return issues, row_issues

    def get_rgb_str(color_obj):
        if color_obj and color_obj.rgb:
            rgb = color_obj.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                return (rgb[-6:] if len(rgb) == 8 else rgb).upper()
        return None

    max_col = inel_sheet.max_column
    total_rows = inel_sheet.max_row

    if show_progress and total_rows > 100:
        print(f"  Checking {total_rows} rows in INEL tab...")

    for row_num in range(2, total_rows + 1):
        if show_progress and total_rows > 100 and row_num % 100 == 0:
            print(f"  Progress: {row_num}/{total_rows} rows checked...", end='\r')

        has_data = False
        for col_num in range(1, max_col + 1):
            cell = inel_sheet.cell(row_num, col_num)
            if cell.value is not None and str(cell.value).strip() != "":
                has_data = True
                break
        if not has_data:
            continue

        has_repeat = False
        repeat_cell = inel_sheet.cell(row_num, max_col)
        if repeat_cell.value:
            cell_text = str(repeat_cell.value).strip().upper()
            if cell_text in ("REPEAT", "LISTED MORE THAN ONCE ON FILE"):
                has_repeat = True

        cells_with_yellow_bg = []
        cells_with_red_font = []
        for col_num in range(1, max_col):
            cell = inel_sheet.cell(row_num, col_num)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            bg_rgb = get_rgb_str(cell.fill.fgColor if cell.fill else None)
            if bg_rgb in ('FFFF00', 'FFFFE0', 'FFFFCC'):
                cells_with_yellow_bg.append((row_num, col_num))
            font_rgb = get_rgb_str(cell.font.color if cell.font else None)
            if font_rgb == 'FF0000':
                cells_with_red_font.append((row_num, col_num))

        if has_repeat:
            repeat_font_ok = repeat_bg_ok = repeat_bold_ok = False
            if repeat_cell.font is not None:
                font_rgb = get_rgb_str(repeat_cell.font.color)
                repeat_font_ok = font_rgb == 'FF0000'
                repeat_bold_ok = bool(repeat_cell.font.bold)
            bg_rgb = get_rgb_str(repeat_cell.fill.fgColor if repeat_cell.fill else None)
            repeat_bg_ok = bg_rgb in ('FFFF00', 'FFFFE0', 'FFFFCC')

            if cells_with_yellow_bg:
                row_issues.append({
                    'row': row_num,
                    'issue_type': 'INEL REPEAT Conflict',
                    'description': f"Row {row_num}: Has 'REPEAT' marker but also has {len(cells_with_yellow_bg)} other highlighted cell(s) - conflicting INEL reasons"
                })

            expected_red_cells = sum(
                1 for c in range(1, max_col)
                if inel_sheet.cell(row_num, c).value is not None
                and str(inel_sheet.cell(row_num, c).value).strip() != ""
            )
            if len(cells_with_red_font) < expected_red_cells:
                row_issues.append({
                    'row': row_num,
                    'issue_type': 'INEL REPEAT Formatting',
                    'description': f"Row {row_num}: REPEAT row should have red font on ALL cells ({len(cells_with_red_font)}/{expected_red_cells} cells have red font)"
                })

            formatting_issues = []
            if not repeat_font_ok:
                formatting_issues.append("red font")
            if not repeat_bold_ok:
                formatting_issues.append("bold")
            if not repeat_bg_ok:
                formatting_issues.append("yellow background")
            if formatting_issues:
                row_issues.append({
                    'row': row_num,
                    'issue_type': 'INEL REPEAT Cell Format',
                    'description': f"Row {row_num}: REPEAT cell missing {', '.join(formatting_issues)}"
                })

        elif not cells_with_yellow_bg:
            row_issues.append({
                'row': row_num,
                'issue_type': 'INEL Missing Reason',
                'description': f"Row {row_num}: No highlighted cells and no REPEAT marker - no indication of why row is in INEL"
            })

    if show_progress and total_rows > 100:
        print()

    return issues, row_issues
