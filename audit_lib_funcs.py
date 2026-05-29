# CAHPS Auditor
# Copyright (C) 2026 HST Pathways. All rights reserved. Developed by Tyler Brock. 
# This copyright notice must be preserved in all copies, modifications, and derivative works of this software.
import re
import datetime
import json
import os
import sys
import csv
from openpyxl.worksheet.worksheet import Worksheet
import phonenumbers
from email_validator import validate_email as ev_validate, EmailNotValidError

# Build-time constants baked in by build_exe.bat; fallback to .env in dev
try:
    from _constants import OAS_SIDS_ONEDRIVE_LINK, HCAHPS_SIDS_ONEDRIVE_LINK  # type: ignore[import]
except ImportError:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv()
    OAS_SIDS_ONEDRIVE_LINK = os.getenv('OAS_SIDS_ONEDRIVE_LINK', '')
    HCAHPS_SIDS_ONEDRIVE_LINK = os.getenv('HCAHPS_SIDS_ONEDRIVE_LINK', '')


# --- SID Registry lookup ---

def _get_sid_file_path(filename):
    """Get the path to a SID registry file from the installation directory."""
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        return os.path.join(exe_dir, filename)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, filename)


def lookup_sid_client_name(sid_prefix, sid_filename='SIDs.csv', onedrive_link=None, show_missing_warning=False, sid_col_idx=0):
    """Look up client name from a SID registry CSV by 2-3 letter SID code.

    Args:
        sid_prefix:        2-3 letter SID code (e.g., 'ANM', 'AGM', 'AC')
        sid_filename:      Name of the SID CSV file to read (default: 'SIDs.csv')
        onedrive_link:     Download URL shown in the missing-file warning (optional)
        show_missing_warning: If True, print warning when the SID file is missing

    Returns:
        Client name string if found, None if not found or error
    """
    if not sid_prefix or len(sid_prefix) < 2 or len(sid_prefix) > 3:
        return None

    csv_path = _get_sid_file_path(sid_filename)

    # Check if file exists
    if not os.path.exists(csv_path):
        if show_missing_warning:
            print("\n" + "="*60)
            print(f"NOTE: {sid_filename} not found")
            print("="*60)
            print(f"The SID registry file ({sid_filename}) is not in the")
            print("installation directory. SID validation will be skipped.")
            if onedrive_link:
                print("")
                print(f"Download {sid_filename} from the shared OneDrive folder:")
                print(f"  {onedrive_link}")
            print("")
            print("Then place it in:")
            print(f"  {os.path.dirname(csv_path)}")
            print("="*60 + "\n")
        return None
    
    def parse_sid_line(raw_line):
        line = raw_line.strip("\r\n")
        if not line.strip():
            return None, None

        # Prefer tab-separated lines (current file format).
        if "\t" in line:
            parts = line.split("\t", 1)
            if sid_col_idx == 0:
                code = parts[0].strip()
                name = parts[1].strip() if len(parts) > 1 else ""
            else:
                name = parts[0].strip()
                code = parts[1].strip() if len(parts) > 1 else ""
            return code, name.strip('"')

        # Fall back to CSV parsing for comma-separated values (legacy format).
        try:
            row = next(csv.reader([line]))
            if len(row) >= 2:
                if sid_col_idx == 0:
                    return row[0].strip(), row[1].strip()
                else:
                    return row[1].strip(), row[0].strip()
        except Exception:
            pass

        # Final fallback: split on any whitespace.
        parts = line.split(None, 1)
        if len(parts) >= 2:
            if sid_col_idx == 0:
                return parts[0].strip(), parts[1].strip().strip('"')
            else:
                return parts[1].strip(), parts[0].strip().strip('"')
        return None, None

    # Read SIDs.csv with a resilient decode strategy so non-UTF8 copies still work.
    encodings_to_try = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
    for enc in encodings_to_try:
        try:
            with open(csv_path, 'r', encoding=enc, errors="strict") as f:
                for raw_line in f:
                    code, name = parse_sid_line(raw_line)
                    if code and code.strip().upper() == sid_prefix.upper():
                        return (name or "").strip()
            # If we successfully parsed the file but did not match, stop.
            return None
        except UnicodeDecodeError:
            continue
        except Exception:
            # Fail silently if CSV can't be read (but exists)
            return None
    
    return None



def get_hf_text(item):
    if item is None:
        return ""
    txt = getattr(item, "text", None)
    if isinstance(txt, str) and txt:
        return txt
    val = getattr(item, "value", None)
    if isinstance(val, str) and val:
        return val
    return str(item) if item else ""


def clean_hf_text(text: str) -> str:
    if not text:
        return ""
    cleaned = re.sub(r"&[A-Z]", "", text)
    cleaned = cleaned.replace("_x000a_", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def pick_header(sheet):
    return (
        get_hf_text(sheet.oddHeader)
        or get_hf_text(sheet.evenHeader)
        or get_hf_text(sheet.firstHeader)
        or ""
    )


def pick_footer(sheet):
    return (
        get_hf_text(sheet.oddFooter)
        or get_hf_text(sheet.evenFooter)
        or get_hf_text(sheet.firstFooter)
        or ""
    )


def normalize_postal_code(raw, state=None):
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "":
        return None
    # Strip non-digits (handles dashes, spaces, ZIP+4 of any form)
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return s  # non-numeric — return as-is so the validator can reject it
    # 8-digit input in a leading-zero state means Excel likely dropped the
    # leading zero from a 9-digit ZIP+4 (e.g. 07123-6789 stored as 71236789)
    if len(digits) == 8 and str(state or "").strip().upper() in _LEADING_ZERO_ZIP_STATES:
        return "0" + digits[:4]
    digits = digits[:5]
    if len(digits) == 4:
        digits = digits.zfill(5)  # restore leading zero dropped by Excel
    return digits


# Known facility, prison, and non-address keywords to flag in ADDRESS1/ADDRESS2
_LEADING_ZERO_ZIP_STATES = {'CT', 'MA', 'ME', 'NH', 'NJ', 'PR', 'RI', 'VT'}
_FACILITY_KEYWORDS = {
    # Correctional / detention facilities
    "lac", "larc", "jail", "prison", "penitentiary", "correctional",
    "detention", "cdcr", "cdc", "fci", "usp", "mdc", "mcc", "cim", "men's central", "century regional", "pitchess",
    "theo lacy", "north county", "central jail", "state prison",
    "county jail", "inmate", "incarcerated",
}

# Non-address placeholders that sometimes appear in ADDRESS1
_PLACEHOLDER_ADDRESSES = {
    "n/a", "na", "none", "unknown", "refused", "decline", "declined",
    "same", "on file", "see above", "no address", "no fixed address",
    "homeless", "transient", "undomiciled", "general delivery",
    "test", "testing", "sample", "tbd", "pending", "null",
}


def check_address(
    sheet,
    street_address_1_col,
    city_col,
    state_col,
    postal_code_col,
    mrn_col=None,
    cms_col=None,
    em_col=None,
    street_address_2_col=None,
    name_col=None,
    age_col=None,
):
    from i18naddress import normalize_address, InvalidAddressError
    import usaddress

    invalid_addresses = []
    noted_addresses = []

    # If any required address column is missing, we can't validate addresses
    if not all([street_address_1_col, city_col, state_col, postal_code_col]):
        return invalid_addresses, noted_addresses

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        mrn  = row[mrn_col  - 1] if mrn_col  else ""
        cms  = row[cms_col  - 1] if cms_col  else ""
        name = str(row[name_col - 1] or "").strip() if name_col else ""
        age  = row[age_col  - 1] if age_col  else ""

        # CMS=2 patients are contacted by email only - skip address checks
        try:
            if int(cms) == 2:
                continue
        except (ValueError, TypeError):
            pass

        em = row[em_col - 1] if em_col else ""

        # E/M=E rows are emailed, not mailed - skip address checks
        if str(em).strip().upper() == "E":
            continue

        street_str = str(row[street_address_1_col - 1] or "").strip()
        street2_str = ""
        if street_address_2_col:
            street2_str = str(row[street_address_2_col - 1] or "").strip()
        city_str = str(row[city_col - 1] or "").strip() or None
        state_str = str(row[state_col - 1] or "").strip() or None
        postal_str = normalize_postal_code(row[postal_code_col - 1], state=state_str)

        # Check for missing fields first
        missing = []
        if not street_str:
            missing.append("street")
        if not city_str:
            missing.append("city")
        if not state_str:
            missing.append("state")
        if not postal_str:
            missing.append("zip")

        if missing:
            invalid_addresses.append(
                f"Row: {row_number} - MRN: '{mrn}' - CMS: '{cms}' - E/M: '{em}' - NAME: '{name}' - AGE: '{age}' - ADDRESS: '{{'street_address': '{street_str}', 'city': '{city_str}', 'country_area': '{state_str}', 'postal_code': '{postal_str}'}}' - REASON: 'Missing: {', '.join(missing)}'"
            )
            continue

        address_data = {
            "country_code": "US",
            "street_address": street_str,
            "city": city_str,
            "country_area": state_str,
            "postal_code": postal_str,
        }

        try:
            normalize_address(address_data)
        except InvalidAddressError as e:
            invalid_addresses.append(
                f"Row: {row_number} - MRN: '{mrn}' - CMS: '{cms}' - E/M: '{em}' - NAME: '{name}' - AGE: '{age}' - ADDRESS: '{address_data}' - REASON: '{e}'"
            )

        # --- Experimental checks (results go into noted_addresses) ---
        note_issues = []

        # 1. Facility / prison keyword check - runs on ALL rows (prisoners should be removed)
        street_lower = street_str.lower()
        street2_lower = street2_str.lower()
        for keyword in _FACILITY_KEYWORDS:
            pattern = rf"\b{re.escape(keyword)}\b"
            match1 = re.search(pattern, street_lower)
            match2 = re.search(pattern, street2_lower) if street2_lower else None
            if match1 or match2:
                field = "ADDRESS1" if match1 else "ADDRESS2"
                note_issues.append(f"Possible facility/institution in {field}: '{keyword}'")
                break  # one match is enough

        # 2-3 only run on mailing rows (E/M = "M")
        is_mailing = str(em).strip().upper() == "M" if em else False

        if is_mailing:
            for placeholder in _PLACEHOLDER_ADDRESSES:
                pattern = rf"^{re.escape(placeholder)}$"
                match1 = re.match(pattern, street_lower)
                match2 = re.match(pattern, street2_lower) if street2_lower else None
                if match1 or match2:
                    field = "ADDRESS1" if match1 else "ADDRESS2"
                    note_issues.append(f"Non-address placeholder in {field}: '{street_str if field == 'ADDRESS1' else street2_str}'")
                    break

            # 3. usaddress structural check - does ADDRESS1 parse as a real street address?
            try:
                tagged, addr_type = usaddress.tag(street_str)
                has_number = "AddressNumber" in tagged
                has_street_name = "StreetName" in tagged or "StreetNamePostType" in tagged
                is_po_box = "USPSBoxType" in tagged

                if addr_type == "Ambiguous":
                    note_issues.append("ADDRESS1 could not be parsed as a street address (ambiguous)")
                elif not is_po_box and not has_number:
                    note_issues.append("ADDRESS1 has no street number")
                elif not is_po_box and not has_street_name:
                    note_issues.append("ADDRESS1 has no street name")
            except usaddress.RepeatedLabelError:  # type: ignore[attr-defined]
                note_issues.append("ADDRESS1 has unusual/repeated address components")

        if note_issues:
            _nc = city_str  or ""
            _ns = state_str or ""
            noted_addresses.append(
                f"Row: {row_number} - MRN: '{mrn}' - CMS: '{cms}' - E/M: '{em}' - NAME: '{name}' - AGE: '{age}' - CITY: '{_nc}' - STATE: '{_ns}' - ADDRESS: '{street_str}' - REASON(s): '{'; '.join(note_issues)}'"
            )
            continue  # skip the city/state/zip-in-street check if we already flagged it

        # Check if city, state, or zip are in the street address field
        if city_str and state_str and postal_str:
            try:
                city_pattern = rf"(?i)(?:(?<=^)|(?<=[\s,])){re.escape(city_str)}(?=(?:,\s*{re.escape(state_str)}|\s+{re.escape(state_str)})(?:\b))"
                state_pattern = rf"(?i)(?:(?<=^)|(?<=[\s,])){re.escape(city_str)}(?=(?:,\s*{re.escape(state_str)}|\s+{re.escape(state_str)})(?:\b)).*?(?:,\s*{re.escape(state_str)}|\s+{re.escape(state_str)})"
                zip_pattern = (
                    rf"(?:(?<=^)|(?<=[\s,])){re.escape(postal_str)}(?:(?=$)|(?=[\s,]))"
                )

                issues = []
                if re.search(city_pattern, street_str):
                    issues.append(city_str)
                if re.search(state_pattern, street_str):
                    issues.append(state_str)
                if re.search(zip_pattern, street_str, re.IGNORECASE):
                    issues.append(postal_str)

                if issues:
                    _nc2 = city_str  or ""
                    _ns2 = state_str or ""
                    noted_addresses.append(
                        f"Row: {row_number} - MRN: '{mrn}' - CMS: '{cms}' - E/M: '{em}' - NAME: '{name}' - AGE: '{age}' - CITY: '{_nc2}' - STATE: '{_ns2}' - ADDRESS: '{street_str}' - REASON(s): '{', '.join(issues)}'"
                    )
            except Exception:
                pass

    return invalid_addresses, noted_addresses


def count_nonempty_rows(sheet):
    """Count rows that actually contain data (ignores blanks/formatting)."""
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            count += 1
    return count


def count_nonempty_rows_after_header(sheet, header_aliases=None):
    """
    Count rows that actually contain data, starting after the header row.
    Finds the header row dynamically (even if not in row 1), then counts
    non-empty rows after it.
    
    Args:
        sheet: The worksheet to count rows in
        header_aliases: List of header column names to search for (defaults to MRN_ALIASES)
    
    Returns:
        int: Count of non-empty data rows after the header
    """
    # Use MRN_ALIASES as default if not provided
    if header_aliases is None:
        # Import here to avoid circular dependency issues
        header_aliases = MRN_ALIASES
    
    # Find the header row by looking for common header column names
    header_row_idx = None
    for row_idx in range(1, 41):  # Check first 40 rows for the header
        try:
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            for cell_value in row:
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if any(cell_str == alias.lower() for alias in header_aliases):
                        header_row_idx = row_idx
                        break
            if header_row_idx:
                break
        except (IndexError, AttributeError):
            continue
    
    # If no header found, default to row 1
    if header_row_idx is None:
        header_row_idx = 1
    
    # Count non-empty rows after the header
    count = 0
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            count += 1
    return count


def is_blank_row(row) -> bool:
    """Return True if the row contains no meaningful (non-empty) values."""
    for cell in row:
        if cell is not None and str(cell).strip() != "":
            return False
    return True


def parse_dob(raw):
    """Parse DOB from various formats (M/D/YYYY, MM/DD/YYYY, YYYY-MM-DD, etc.).
    Returns (ok, normalized, error_reason). Normalized is always MM/DD/YYYY when ok.
    Flags dates that are invalid, more than 120 years in the past, or in the future.
    """
    if raw is None:
        return False, None, "blank"

    s = str(raw).strip()
    if s.startswith("'"):
        s = s[1:].strip()

    # Try common date formats
    formats = [
        "%m/%d/%Y",  # 3/10/1986 or 03/10/1986
        "%Y-%m-%d",  # 1986-03-10
        "%m-%d-%Y",  # 03-10-1986
    ]

    # Handle datetime objects or strings with time components
    if " " in s:
        s = s.split()[0]  # Take just the date part

    dt = None
    for fmt in formats:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            break
        except ValueError:
            continue

    if dt is None:
        return False, None, "invalid date format"

    # Check if date is valid and within reasonable range
    now = datetime.datetime.now()
    years_ago_120 = now - datetime.timedelta(days=120 * 365.25)

    if dt > now:
        return False, None, "future date"
    if dt < years_ago_120:
        return False, None, "more than 120 years old"

    # Return normalized format MM/DD/YYYY
    return True, dt.strftime("%m/%d/%Y"), None


def validate_sid_sequence(sheet, sid_col, cms_col, header_sid=None):
    """
    Validate SID sequence for proper formatting, uniqueness, and numerical order.
    Only validates rows where CMS INDICATOR = 1.
    Returns (issues, row_issues) lists.
    
    Args:
        sheet: The worksheet to validate
        sid_col: Column index for SID (1-based), or None if column missing
        cms_col: Column index for CMS INDICATOR (1-based), or None if column missing
        header_sid: The SID from the header (should be first SID - 1)
    """
    issues = []
    row_issues = []
    
    # Return empty results if required columns are missing
    if sid_col is None or cms_col is None:
        return issues, row_issues
    
    sid_pattern = re.compile(r'^([A-Z]{2,3})(\d+)$')
    
    sids_found = []
    expected_prefix = None
    expected_start_num = None
    cms1_rows_processed = 0
    first_sid_encountered = False
    
    if header_sid:
        header_match = sid_pattern.match(str(header_sid).strip().upper())
        if header_match:
            expected_prefix = header_match.group(1)
            expected_start_num = int(header_match.group(2)) + 1
        else:
            issues.append(f"Header SID '{header_sid}' does not match expected format (3 letters + numbers)")
    
    row_num = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell for cell in row):
            break
            
        cms_value = row[cms_col - 1] if cms_col <= len(row) else None
        
        try:
            cms_int = int(float(str(cms_value).strip())) if cms_value is not None and str(cms_value).strip() != "" else None
        except (ValueError, TypeError):
            cms_int = None
        
        if cms_int != 1:
            # Check if a SID was accidentally entered on a non-CMS=1 row
            sid_value_check = row[sid_col - 1] if sid_col <= len(row) else None
            if sid_value_check is not None and str(sid_value_check).strip():
                mrn_value_check = row[0] if len(row) > 0 else None
                row_issues.append({
                    'row': row_num,
                    'mrn': mrn_value_check,
                    'cms': cms_value,
                    'issue_type': 'SID on Non-CMS=1 Row',
                    'description': f"Row {row_num}: SID '{str(sid_value_check).strip()}' found on row with CMS={cms_value} (expected CMS=1 only)"
                })
            row_num += 1
            continue
            
        cms1_rows_processed += 1
        sid_value = row[sid_col - 1] if sid_col <= len(row) else None
        mrn_value = row[0] if len(row) > 0 else None
        
        if sid_value is None or str(sid_value).strip() == "":
            row_issues.append({
                'row': row_num,
                'mrn': mrn_value,
                'cms': cms_value,
                'issue_type': 'SID Missing',
                'description': f"Row {row_num}: SID is missing or empty (CMS=1)"
            })
            row_num += 1
            continue
            
        sid_str = str(sid_value).strip().upper()
        
        match = sid_pattern.match(sid_str)
        if not match:
            row_issues.append({
                'row': row_num,
                'mrn': mrn_value,
                'cms': cms_value,
                'issue_type': 'SID Format',
                'description': f"Row {row_num}: SID '{sid_str}' does not match format (3 letters + numbers)"
            })
            row_num += 1
            continue
            
        prefix = match.group(1)
        number = int(match.group(2))
        if not first_sid_encountered:
            first_sid_encountered = True
            if expected_prefix is None:
                expected_prefix = prefix
            if expected_start_num is None:
                expected_start_num = number
                expected_start_num = number
        
        if prefix != expected_prefix:
            row_issues.append({
                'row': row_num,
                'mrn': mrn_value,
                'cms': cms_value,
                'issue_type': 'SID Prefix',
                'description': f"Row {row_num}: SID prefix '{prefix}' does not match expected '{expected_prefix}'"
            })
        
        if sid_str in sids_found:
            row_issues.append({
                'row': row_num,
                'mrn': mrn_value,
                'cms': cms_value,
                'issue_type': 'SID Duplicate',
                'description': f"Row {row_num}: Duplicate SID '{sid_str}'"
            })
        
        sids_found.append(sid_str)
        
        if expected_start_num is not None:
            expected_num = expected_start_num + (cms1_rows_processed - 1)
            if number != expected_num:
                row_issues.append({
                    'row': row_num,
                    'mrn': mrn_value,
                    'cms': cms_value,
                    'issue_type': 'SID Sequence',
                    'description': f"Row {row_num}: Expected SID '{expected_prefix}{expected_num:05d}', found '{sid_str}'"
                })
        
        row_num += 1
    
    if row_issues:
        issues.append(f"Found {len(row_issues)} SID validation issues")
    
    return issues, row_issues


# validate_inel_repeat_rows has been moved to audit_oas_funcs.py (OAS-specific)


# --- Cross-tab consistency checking ---

# MRN and Email alias mappings (from VBA script)
MRN_ALIASES = [
    "chart id",
    "patid",
    "medical account number",
    "patient account number",
    "patient acct no",
    "medical record number",
    "medical_record_number",
    "mrn",
    "patient id",
    "patient mrn",
    "medicalrecordnumber",
    "medrec",
    "md rc",
    "acct#",
    "patient account #",
    "account number",
    "patient chart number",
    "acctnum",
    "mrnum",
    "patientid",
    "mrno",
    "per nbr",
    "pt.id",
    "chart number",
    "pt account #",
    "mr#",
    "patient_number",
    "pat_med_rec",
    "person mrn",
    "armrnum",
    "med rec number",
    "person_nbr",
    "pt id #",
    "armrnum-t",
    "emr number",
    "patient - patient - id",
    "pt_id",
    "mrn #",
    "chart no.",
    "v#",
    "pat person nbr",
    "med rec #",
    "patient - id",
    "med_rec_nbr",
]

EMAIL_ALIASES = [
    "e-mail address",
    "emailaddress",
    "email",
    "email address",
    "patient email",
    "e-mail",
    "patientemailaddress",
    "patient e-mail",
    "pm_email",
    "patient email address",
    "patemail",
    "patient email address",
    "email addr",
    "pt. email id",
    "email addr",
    "pt email",
    "patmail",
    "pt_email_address",
    "per_email:per addr street 1",
    "arpatmail",
    "per email (addr type)",
    "patient email address",
    "arsubmail-t",
    "arsubmail",
    "patient - email",
    "email id",
    "pat_email",
    "patientaddressemail",
]

SERVICE_DATE_ALIASES = [
    "proc:proc dt/tm",
    "date of procedure",
    "patient procedure date",
    "date time",
    "dos",
    "service date",
    "date of service",
    "dateofprocedure",
    "adm_svc_date",
    "dt of svc",
    "patient hospital procedure date",
    "cpt date",
    "procedure date",
    "date of patient procedure date",
    "appointment date",
    "patient asc procedure date",
    "srvdate",
    "srvday",
    "os date",
    "service_date_from",
    "patient hospital procedure ...",
    "charge date of service",
    "surgery date",
    "dateofprocedure",
    "proc_date",
    "procedure_date",
    "admdt",
    "patvisitdt date",
    "date of surgery",
    "visit or admit date",
    "admission date",
    "visit date",
    "appt date",
    "visitoradmitdate",
    "encounter date",
    "admit date",
    "date of admission",
    "admission",
    "enc registration dt/tm",
    "case date",
    "patient hospital discharge date",
    "aradmdt",
    "arprdt01",
    "chargeservicedate",
    "procedure dos",
    "patient procdure date",
    "aradmdt-n",
    "procedure - dos",
    "date",
    "proc date 01",
    "patient asc procdure date",
    "procedure date",
    "appt_datetime",
    "dos from date",
    "servdate",
    "srvc date",
    "date svc from",
    "patient hosp procedure date",
    "appt dt",
    "scheduledate",
    "clmcreatedday",
    "from dos",
    "case dos",
    "d.date",
]

MRN_ALIASES = [
    "chart id", "patid", "medical account number", "patient account number",
    "patient acct no", "medical record number", "mrn", "patient id",
    "patient mrn", "medicalrecordnumber", "medrec", "md rc", "acct#",
    "patient account #", "account number", "patient chart number",
    "acctnum", "mrnum", "patientid", "mrno", "per nbr", "pt.id",
    "chart number", "pt account #", "mr#", "patient_number",
    "pat_med_rec", "person mrn", "armrnum", "med rec number",
    "person_nbr", "pt id #", "armrnum-t", "emr number",
    "patient - patient - id", "pt_id", "mrn #", "chart no.",
    "v#", "pat person nbr", "med rec #", "patient - id",
    "med_rec_nbr", "patient chart id",
]

FACILITY_NAME_ALIASES = [
    "facility name",
    "facility",
    "facility_name",
    "facilityname",
    "fac name",
    "fac_name",
    "facname",
    "facility nm",
    "facility_nm",
    "fac nm",
    "site name",
    "site_name",
    "sitename",
    "site",
    "location name",
    "location_name",
    "locationname",
    "location",
    "clinic name",
    "clinic_name",
    "clinicname",
    "clinic",
    "hospital name",
    "hospital_name",
    "hospitalname",
    "hospital",
    "center name",
    "center_name",
    "centername",
    "practice name",
    "practice_name",
    "practicename",
    "practice",
    "asc name",
    "asc_name",
    "ascname",
    "surgery center",
    "surgery_center",
    "surgerycenter",
    "surgical center",
    "surgical_center",
    "surgicalcenter",
    "vendor name",
    "vendor_name",
    "vendorname",
    "vendor",
    "org name",
    "org_name",
    "orgname",
    "organization name",
    "organization_name",
    "organizationname",
    "organization",
    "client name",
    "client_name",
    "clientname",
    "client",
    "entity name",
    "entity_name",
    "entityname",
    "fac",
    "loc",
    "loc name",
    "loc_name",
    "agency name",
    "agency_name",
    "agencyname",
    "agency",
    "client id",
    "client_id",
    "clientid",
    "revenue location",
    "revenue_location",
    "revenuelocation",
]

# Headers that often mark the start of patient-level data blocks in POP-like layouts.
# If these appear while collecting facility/location values, we stop collecting.
_DATA_BLOCK_HEADER_MARKERS = {
    "patient first name",
    "first name",
    "patientfirstname",
    "firstname",
    "patient last name",
    "last name",
    "patientlastname",
    "lastname",
    "patient name",
    "name",
    "mrn",
    "medical record number",
    "chart id",
    "patient id",
    "dob",
    "date of birth",
    "gender",
    "service date",
    "email",
    "email address",
    "address",
    "address1",
    "address2",
    "city",
    "state",
    "zip",
}

_DATA_HEADER_KEYWORD_HINTS = (
    "patient",
    "mrn",
    "chart",
    "id",
    "dob",
    "birth",
    "gender",
    "email",
    "phone",
    "telephone",
    "date",
    "address",
    "city",
    "state",
    "zip",
)


def _expand_alias_variants(alias):
    """Generate matching variants of an alias. Mirrors VBA ExpandAliases logic:
    original, underscores-to-spaces, remove underscores, remove spaces,
    spaces-to-underscores, spaces-to-dashes. All lowercase."""
    low = alias.strip().lower()
    variants = {low}
    variants.add(low.replace("_", " "))   # underscores -> spaces
    variants.add(low.replace("_", ""))    # remove underscores
    variants.add(low.replace(" ", ""))    # remove spaces
    variants.add(low.replace(" ", "_"))   # spaces -> underscores
    variants.add(low.replace(" ", "-"))   # spaces -> dashes
    return variants


def _expand_aliases(aliases):
    """Expand a list of aliases into a set including all space/underscore variants."""
    expanded = set()
    for alias in aliases:
        expanded.update(_expand_alias_variants(alias))
    return expanded


def _is_likely_data_block_header(value):
    """Return True when a value looks like a patient-data header label."""
    if value is None:
        return False
    normalized = re.sub(r'\s+', ' ', str(value)).strip().strip('"\'').strip().lower()
    compact = normalized.replace(" ", "").replace("_", "").replace("-", "")
    if normalized in _DATA_BLOCK_HEADER_MARKERS or compact in _DATA_BLOCK_HEADER_MARKERS:
        return True
    # Use word-boundary matching on the space-preserving normalized form so that
    # short keywords like "id" don't fire on unrelated words ("provider", "valid", etc.)
    return any(re.search(r'(?<![a-z])' + re.escape(k) + r'(?![a-z])', normalized) for k in _DATA_HEADER_KEYWORD_HINTS)


def _row_has_data_header_signature(values, min_hits=3):
    """Heuristic: row looks like a patient-data header if several cells are header-like."""
    hits = 0
    for cell in values:
        if cell is None:
            continue
        txt = str(cell).strip()
        if not txt:
            continue
        if _is_likely_data_block_header(txt):
            hits += 1
            if hits >= min_hits:
                return True
    return False


def _detect_sheet_delimiter(sheet, check_rows=15):
    """
    Detect if a sheet has data packed into a single column using | or , as a delimiter.
    This handles POP tabs where all columns are joined into one cell per row.

    Returns (delimiter, header_row_idx, header_parts) or (None, None, None) if normal.
    Pipe (|) is checked before comma to avoid false positives on data cells that
    legitimately contain commas (e.g. CPT code strings like "43239,FAC").
    """
    for row_idx in range(1, check_rows + 1):
        try:
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        except (IndexError, AttributeError):
            continue

        non_empty = [c for c in row if c is not None and str(c).strip()]
        if not non_empty:
            continue

        # Only attempt detection when very few real columns are populated (1-3)
        if len(non_empty) <= 3:
            cell_str = str(non_empty[0]).strip()
            for delim in ['|', ',']:
                if delim in cell_str:
                    parts = [p.strip() for p in cell_str.split(delim)]
                    # Require at least 4 parts so we don't misfire on simple data cells
                    if len(parts) >= 4:
                        return delim, row_idx, parts

    return None, None, None


def find_column_in_sheet(sheet, aliases):
    """
    Like find_column_by_aliases but also handles pipe/comma delimited single-column sheets.
    Returns a dict with column info, or None if not found:
      {
        'col_idx':    int  - 1-based for normal sheets; 0-based position within the split for delimited
        'header_row': int  - row index where the header was found
        'delimiter':  str or None
        'is_delimited': bool
        'header_name': str
      }
    """
    # Try normal column layout first
    col_idx, hdr_row = find_column_by_aliases(sheet, aliases)
    if col_idx is not None:
        try:
            hdr_cells = list(sheet.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True))[0]
            hdr_name = str(hdr_cells[col_idx - 1]).strip()
        except Exception:
            hdr_name = ''
        return {
            'col_idx': col_idx,
            'header_row': hdr_row,
            'delimiter': None,
            'is_delimited': False,
            'header_name': hdr_name,
        }

    # Fall back to delimited detection
    delimiter, hdr_row, parts = _detect_sheet_delimiter(sheet)
    if delimiter is None:
        return None

    expanded = _expand_aliases(aliases)
    for pos, part in enumerate(parts or []):
        cell_str = re.sub(r'\s+', ' ', part).strip().strip('"\'').strip().lower()
        if cell_str in expanded:
            return {
                'col_idx': pos,   # 0-based position within the split
                'header_row': hdr_row,
                'delimiter': delimiter,
                'is_delimited': True,
                'header_name': part,
            }

    return None


def get_row_value(row, col_info):
    """
    Extract the value for a column from a data row, handling delimited sheets.
    col_info: dict returned by find_column_in_sheet.
    """
    if col_info is None:
        return None
    if col_info['is_delimited']:
        cell = row[0] if row else None
        if cell is None:
            return None
        parts = str(cell).split(col_info['delimiter'])
        idx = col_info['col_idx']
        return parts[idx].strip() if idx < len(parts) else None
    else:
        idx = col_info['col_idx'] - 1
        return row[idx] if idx < len(row) else None


def find_column_by_aliases(sheet, aliases):
    """
    Find a column in the sheet by checking against a list of aliases.
    Returns the 1-based column index if found, None otherwise.
    Handles sheets with rows spaced apart.
    Also checks underscore/space/removed variants of each alias.
    """
    expanded = _expand_aliases(aliases)
    # Check first few rows for headers (in case of spacing)
    for header_row_idx in range(1, 41):  # Check first 40 rows
        try:
            row = list(
                sheet.iter_rows(
                    min_row=header_row_idx, max_row=header_row_idx, values_only=True
                )
            )[0]
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if cell_str in expanded:
                        return col_idx, header_row_idx
        except (IndexError, AttributeError):
            continue
    return None, None


def find_all_columns_by_aliases(sheet, aliases):
    """
    Find ALL columns in the sheet that match any alias in the list.
    Returns a list of dicts: [{'col': 1-based col index, 'header_row': row index,
                               'header_name': original header text, 'values': list of unique non-empty values}]
    """
    expanded = _expand_aliases(aliases)
    found = []
    seen_cols = set()
    for header_row_idx in range(1, 41):  # Check first 40 rows
        try:
            row = list(
                sheet.iter_rows(
                    min_row=header_row_idx, max_row=header_row_idx, values_only=True
                )
            )[0]
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value and col_idx not in seen_cols:
                    # Normalize: collapse all whitespace, strip quotes, lowercase
                    cell_str = re.sub(r'\s+', ' ', str(cell_value)).strip().strip('"\'').strip().lower()
                    if cell_str in expanded:
                        # Collect all unique non-empty values from this column
                        unique_vals = []
                        seen_vals = set()
                        for data_row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                            val = data_row[col_idx - 1] if len(data_row) >= col_idx else None
                            if unique_vals and _row_has_data_header_signature(data_row):
                                break
                            if val is not None and str(val).strip():
                                if unique_vals and _is_likely_data_block_header(val):
                                    break
                                val_str = str(val).strip()
                                val_lower = val_str.lower()
                                if val_lower not in seen_vals:
                                    seen_vals.add(val_lower)
                                    unique_vals.append(val_str)
                        found.append({
                            'col': col_idx,
                            'header_row': header_row_idx,
                            'header_name': str(cell_value).strip(),
                            'values': unique_vals,
                        })
                        seen_cols.add(col_idx)
        except (IndexError, AttributeError):
            continue
    return found


def find_all_columns_in_sheet(sheet, aliases):
    """
    Like find_all_columns_by_aliases but also handles pipe/comma delimited single-column sheets.
    Returns a list of match dicts (header_name, values, col_idx, header_row, is_delimited, delimiter).
    Tries normal layout first; if nothing found, falls back to delimiter detection.
    """
    results = find_all_columns_by_aliases(sheet, aliases)
    if results:
        for r in results:
            r['is_delimited'] = False
            r['delimiter'] = None
        return results

    # Fall back to delimiter detection
    delimiter, hdr_row, parts = _detect_sheet_delimiter(sheet)
    if delimiter is None:
        return []

    expanded = _expand_aliases(aliases)
    found = []
    for pos, part in enumerate(parts or []):
        cell_str = re.sub(r'\s+', ' ', part).strip().strip('"\'').strip().lower()
        if cell_str in expanded:
            unique_vals = []
            seen_vals = set()
            for data_row in sheet.iter_rows(min_row=(hdr_row or 0) + 1, values_only=True):
                raw = data_row[0] if data_row else None
                if raw is None:
                    continue
                row_parts = str(raw).split(delimiter)
                if unique_vals and _row_has_data_header_signature(row_parts):
                    break
                val = row_parts[pos].strip() if pos < len(row_parts) else ''
                if val:
                    if unique_vals and _is_likely_data_block_header(val):
                        break
                    if val.lower() in seen_vals:
                        continue
                    seen_vals.add(val.lower())
                    unique_vals.append(val)
            found.append({
                'col_idx': pos,
                'header_row': hdr_row,
                'header_name': part,
                'values': unique_vals,
                'is_delimited': True,
                'delimiter': delimiter,
            })
    return found


def normalize_email(email_val):
    """Normalize email for comparison (lowercase, stripped)."""
    if email_val is None:
        return None
    email_str = str(email_val).strip().lower()
    if email_str == "" or email_str == "none":
        return None
    return email_str


def check_pop_upload_email_consistency(
    wb, upload_sheet, mrn_col_upload, email_col_upload
):
    """
    Check that emails in UPLOAD tab match those in POP tab for the same MRN.
    Returns list of mismatches: [(upload_row, mrn, upload_email, pop_email), ...]
    """
    mismatches = []

    # Check if POP tab exists
    if "POP" not in wb.sheetnames:
        return mismatches  # Can't check without POP tab

    pop_sheet = wb["POP"]

    # Find MRN and Email columns in POP using aliases - handles both normal and delimited sheets
    mrn_info = find_column_in_sheet(pop_sheet, MRN_ALIASES)
    email_info = find_column_in_sheet(pop_sheet, EMAIL_ALIASES)

    if mrn_info is None:
        return [("N/A", "N/A", "N/A", "Could not locate MRN column in POP tab")]

    if email_info is None:
        return [("N/A", "N/A", "N/A", "Could not locate Email column in POP tab")]

    # Build a dictionary of MRN -> Email from POP tab
    pop_data_start_row = mrn_info['header_row'] + 1

    pop_mrn_to_email = {}
    for row in pop_sheet.iter_rows(min_row=pop_data_start_row, values_only=True):
        # Skip completely empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        # Get MRN and Email from this row (works for both normal and delimited sheets)
        try:
            mrn_val = get_row_value(row, mrn_info)
            email_val = get_row_value(row, email_info)

            if mrn_val:
                mrn_str = str(mrn_val).strip()
                if mrn_str:
                    # Store normalized email
                    pop_mrn_to_email[mrn_str] = normalize_email(email_val)
        except (IndexError, AttributeError):
            continue

    # Now compare UPLOAD tab against POP data
    for upload_row_idx, row in enumerate(
        upload_sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        try:
            upload_mrn = row[mrn_col_upload - 1] if mrn_col_upload <= len(row) else None
            upload_email = (
                row[email_col_upload - 1] if email_col_upload <= len(row) else None
            )

            if not upload_mrn:
                continue

            mrn_str = str(upload_mrn).strip()
            upload_email_norm = normalize_email(upload_email)

            # Check if this MRN exists in POP
            if mrn_str in pop_mrn_to_email:
                pop_email_norm = pop_mrn_to_email[mrn_str]

                # Compare emails (only flag if both exist and differ)
                if upload_email_norm and pop_email_norm:
                    if upload_email_norm != pop_email_norm:
                        mismatches.append(
                            (
                                upload_row_idx,
                                mrn_str,
                                upload_email or "",
                                pop_mrn_to_email[mrn_str] or "",
                            )
                        )
        except (IndexError, AttributeError):
            continue

    return mismatches


def extract_service_date_range(sheet, svc_col, mrn_col=None, cms_col=None):
    """
    Extract the earliest and latest service dates from the SERVICE DATE column.
    Also validates that no SERVICE DATE fields are blank.
    
    Returns: (date_range_str, blank_date_issues, blank_date_row_issues)
        - date_range_str: "MM/DD/YYYY - MM/DD/YYYY" or None if no valid dates
        - blank_date_issues: List of general issue strings
        - blank_date_row_issues: List of dicts with row-level blank date issues
    """
    blank_date_issues = []
    blank_date_row_issues = []
    
    if svc_col is None:
        return None, blank_date_issues, blank_date_row_issues
    
    valid_dates = []
    
    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if is_blank_row(row):
            continue
        
        mrn_val = row[mrn_col - 1] if mrn_col and mrn_col <= len(row) else None
        cms_val = row[cms_col - 1] if cms_col and cms_col <= len(row) else None
        svc_val = row[svc_col - 1] if svc_col <= len(row) else None
        
        # Check for blank SERVICE DATE
        if svc_val is None or str(svc_val).strip() == "":
            blank_date_row_issues.append({
                'row': r,
                'mrn': mrn_val,
                'cms': cms_val,
                'issue_type': 'Blank Service Date',
                'description': f"SERVICE DATE is blank or empty"
            })
            blank_date_issues.append(f"OASCAPHS Row {r}: SERVICE DATE is blank")
            continue
        
        # Try to parse the date
        try:
            if isinstance(svc_val, datetime.datetime):
                valid_dates.append(svc_val)
            else:
                svc_str = str(svc_val).strip()
                # Try to parse MM/DD/YYYY format
                date_obj = datetime.datetime.strptime(svc_str, "%m/%d/%Y")
                valid_dates.append(date_obj)
        except (ValueError, AttributeError):
            # Invalid date format - this will be caught by column_validations
            pass
    
    # Return date range if we have valid dates
    if valid_dates:
        earliest = min(valid_dates)
        latest = max(valid_dates)
        date_range_str = f"{earliest.strftime('%m/%d/%Y')} - {latest.strftime('%m/%d/%Y')}"
        return date_range_str, blank_date_issues, blank_date_row_issues
    
    return None, blank_date_issues, blank_date_row_issues


def column_validations(sheet, headers, mrn_col, cms_col, em_col, issues, row_issues, filename_year=None):
    """
    Perform data quality validation checks on OASCAPHS sheet columns.
    Returns updated issues and row_issues lists.
    """
    from collections import defaultdict

    svc_col = headers.get("SERVICE DATE") or headers.get("D.DATE")
    age_col = headers.get("AGE")
    email_col = headers.get("EMAIL ADDRESS")
    lang_col = headers.get("SURVEY LANGUAGE")
    tel_col = headers.get("TELEPHONE")
    dob_col = headers.get("DATE OF BIRTH")
    name_col = headers.get("PATIENT NAME")
    gender_col = headers.get("GENDER")

    # Track service dates to check they're all in the same month
    service_dates = []
    # Track MRNs to check for duplicates
    mrn_tracker = defaultdict(list)

    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if is_blank_row(row):
            continue

        mrn_val = row[mrn_col - 1] if mrn_col else None
        cms_val = row[cms_col - 1] if cms_col else None
        em_val = row[em_col - 1] if em_col else None

        # Track MRN for duplicate check
        if mrn_val:
            mrn_tracker[mrn_val].append(r)

        # GENDER - must be M, F, 0, 1, or 2 (blank is acceptable)
        if gender_col:
            gender_val = row[gender_col - 1]
            valid_genders = ["M", "F", "0", "1", "2", "U", "O"]
            gender_str = str(gender_val).strip().upper() if gender_val else ""
            if gender_str and gender_str not in valid_genders:
                row_issues.append(
                    {
                        "row": r,
                        "mrn": mrn_val,
                        "cms": cms_val,
                        "issue_type": "Invalid Gender",
                        "description": f"Gender '{gender_val}' not in {valid_genders}",
                    }
                )

        # SERVICE DATE - validate format and collect all dates for month validation
        if svc_col:
            svc_val = row[svc_col - 1]
            if svc_val:
                # Convert to string for validation
                if isinstance(svc_val, datetime.datetime):
                    svc_str = svc_val.strftime("%m/%d/%Y")
                    service_dates.append((r, mrn_val, svc_val))
                else:
                    svc_str = str(svc_val).strip()

                # Check format MM/DD/YYYY
                if not re.match(
                    r"^(0[1-9]|1[0-2])/(0[1-9]|[12][0-9]|3[01])/\d{4}$", svc_str
                ):
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": "Invalid Service Date Format",
                            "description": f"Service Date '{svc_str}' must be MM/DD/YYYY format",
                        }
                    )
                    continue

                # Check if date is in the future
                try:
                    svc_date = datetime.datetime.strptime(svc_str, "%m/%d/%Y")
                    if svc_date > datetime.datetime.now():
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Service Date In Future",
                                "description": f"Service Date '{svc_str}' is in the future",
                            }
                        )
                    else:
                        # Only add valid dates for month validation
                        service_dates.append((r, mrn_val, svc_date))
                except ValueError:
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": "Invalid Service Date",
                            "description": f"Service Date '{svc_str}' is not a valid date",
                        }
                    )

        # AGE - must be 18 or older (only matters when CMS=1)
        if age_col:
            age_val = row[age_col - 1]
            try:
                age_int = int(float(str(age_val))) if age_val is not None else None
                cms_int = (
                    int(float(str(cms_val)))
                    if cms_val is not None and str(cms_val).strip()
                    else None
                )

                if age_int is not None and cms_int == 1:
                    if age_int <= 0:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Invalid Age",
                                "description": f"Age {age_int} is not a valid age (CMS=1)",
                            }
                        )
                    elif age_int < 18:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Age Too Young",
                                "description": f"Age {age_int} is below 18 (CMS=1)",
                            }
                        )
                    elif age_int > 110:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Age Suspicious",
                                "description": f"Age {age_int} is implausibly high (CMS=1)",
                            }
                        )
            except (ValueError, TypeError):
                pass

        # make sure date of birth is valid (day, month, and year are present and not in the future). it should look exactly like this: 01/01/2025, for example
        if dob_col:
            dob_val = row[dob_col - 1]
            if dob_val:
                ok, normalized, err = parse_dob(dob_val)
                if not ok:
                    issue_type = "DOB In Future" if err == "future" else "Invalid DOB"
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": issue_type,
                            "description": f"DOB '{dob_val}' error: {err}",
                        }
                    )

        # EMAIL ADDRESS - validate format when present; require it for CMS=2
        if email_col:
            email_val = row[email_col - 1]
            if email_val and str(email_val).strip():
                email_str = str(email_val).strip()
                # Use email-validator for RFC-compliant syntax checking (no DNS)
                try:
                    ev_validate(email_str, check_deliverability=False)
                except EmailNotValidError as e:
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": "Invalid Email Format",
                            "description": f"Email '{email_str}' - {e}",
                        }
                    )
            else:
                # CMS=2 patients are email-only - a missing email means they can't be contacted
                try:
                    if cms_val is not None and int(cms_val) == 2:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Missing Email for CMS=2",
                                "description": "CMS=2 but email address is blank (email is the only contact method)",
                            }
                        )
                except (ValueError, TypeError):
                    pass
                # E/M=E rows are sent via email - a missing email means they won't receive a survey
                em_str = str(em_val).strip().upper() if em_val else ""
                if em_str == "E":
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": "Missing Email for E/M=E",
                            "description": "E/M is 'E' but email address is blank",
                        }
                    )

        # SURVEY LANGUAGE - must be en, es, ko, zh, or m (lowercase)
        if lang_col:
            lang_val = row[lang_col - 1]
            valid_langs = ["en", "es", "ko", "zh", "m"]
            lang_str = str(lang_val).strip() if lang_val else ""
            if not lang_str or lang_str not in valid_langs:
                row_issues.append(
                    {
                        "row": r,
                        "mrn": mrn_val,
                        "cms": cms_val,
                        "issue_type": "Invalid Language Code",
                        "description": f"Language '{lang_str}' not in {valid_langs}",
                    }
                )

        # E/M and CMS INDICATOR logic
        # - If CMS=1, E/M must be 'E' or 'M'
        # - If CMS=2, E/M should NOT be 'E' or 'M'
        if cms_col and em_col:
            try:
                cms_int = (
                    int(float(str(cms_val)))
                    if cms_val is not None and str(cms_val).strip()
                    else None
                )
                em_str = str(em_val).strip().upper() if em_val else ""

                if cms_int == 1:
                    if em_str not in ["E", "M"]:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Missing E/M for CMS=1",
                                "description": f"CMS=1 but E/M is '{em_val}' (expected 'E' or 'M')",
                            }
                        )
                elif cms_int == 2:
                    if em_str in ["E", "M"]:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Unexpected E/M for CMS=2",
                                "description": f"CMS=2 but E/M is '{em_val}' (should be blank)",
                            }
                        )
            except (ValueError, TypeError):
                pass

    # Check all SERVICE DATEs are in the same month
    if service_dates:
        # Get month/year from first date
        first_date = service_dates[0][2]
        expected_month = first_date.month
        expected_year = first_date.year

        for r, mrn_val, svc_date in service_dates:
            if svc_date.month != expected_month or svc_date.year != expected_year:
                row_issues.append(
                    {
                        "row": r,
                        "mrn": mrn_val,
                        "cms": None,
                        "issue_type": "Service Date Wrong Month",
                        "description": f"Date {svc_date.strftime('%Y-%m-%d')} not in {expected_year}-{expected_month:02d}",
                    }
                )

    # Check all SERVICE DATEs are in the year from the filename
    if filename_year is not None and service_dates:
        for r, mrn_val, svc_date in service_dates:
            if svc_date.year != filename_year:
                row_issues.append(
                    {
                        "row": r,
                        "mrn": mrn_val,
                        "cms": None,
                        "issue_type": "Service Date Wrong Year",
                        "description": f"Date {svc_date.strftime('%m/%d/%Y')} is not in {filename_year} (filename year)",
                    }
                )

    # Check for duplicate MRNs
    for mrn, rows in mrn_tracker.items():
        if len(rows) > 1:
            rows_str = ", ".join(str(r) for r in rows)
            for r in rows:
                row_issues.append(
                    {
                        "row": r,
                        "mrn": mrn,
                        "cms": None,
                        "issue_type": "Duplicate MRN",
                        "description": f"MRN appears in rows: {rows_str}",
                    }
                )
            issues.append(f"Duplicate MRN '{mrn}' found in rows {rows_str}")

    # check validity of telephone numbers using phonenumbers package
    if tel_col:
        for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if is_blank_row(row):
                continue
            tel_val = row[tel_col - 1]
            mrn_val = row[mrn_col - 1] if mrn_col else None
            cms_val = row[cms_col - 1] if cms_col else None

            # CMS=2 patients are contacted by email only - skip phone checks
            try:
                if cms_val is not None and int(cms_val) == 2:
                    continue
            except (ValueError, TypeError):
                pass

            if tel_val and str(tel_val).strip():
                tel_str = str(tel_val).strip()
                try:
                    phone_number = phonenumbers.parse(tel_str, "US")
                    if not phonenumbers.is_valid_number(phone_number):
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Invalid Telephone Number",
                                "description": f"Telephone '{tel_str}' is not a valid number",
                            }
                        )
                except phonenumbers.NumberParseException:
                    row_issues.append(
                        {
                            "row": r,
                            "mrn": mrn_val,
                            "cms": cms_val,
                            "issue_type": "Invalid Telephone Number Format",
                            "description": f"Telephone '{tel_str}' has invalid format",
                        }
                    )

    # Check for duplicate phone numbers (possible accidental copy-paste)
    if tel_col:
        phone_tracker = defaultdict(list)
        for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if is_blank_row(row):
                continue
            tel_val = row[tel_col - 1]
            mrn_val = row[mrn_col - 1] if mrn_col else None
            cms_val = row[cms_col - 1] if cms_col else None
            if tel_val and str(tel_val).strip():
                phone_tracker[str(tel_val).strip()].append((r, mrn_val, cms_val))
        for tel_str, entries in phone_tracker.items():
            if len(entries) > 1:
                # If cms_col is present, only flag when at least 2 rows are CMS=1
                # (avoids false positives for non-reported patients sharing a number)
                if cms_col:
                    cms1_appearances = sum(
                        1 for _, _, cms_val in entries
                        if cms_val is not None and str(cms_val).strip() == "1"
                    )
                    if cms1_appearances < 2:
                        continue
                rows_str = ", ".join(str(e[0]) for e in entries)
                for r, mrn_val, cms_val in entries:
                    row_issues.append({
                        "row": r,
                        "mrn": mrn_val,
                        "cms": cms_val,
                        "issue_type": "Duplicate Telephone Number",
                        "description": f"Phone '{tel_str}' appears in rows: {rows_str}",
                    })
                issues.append(f"Duplicate phone '{tel_str}' found in rows {rows_str}")

    # find placeholder/test names in patient name col
    if name_col:
        placeholder_names = {
            "test",
            "patient",
            "sample",
            "john doe",
            "jane doe",
            "asdf",
            "qwerty",
            "foo bar",
        }
        for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if is_blank_row(row):
                continue
            name_val = row[name_col - 1]
            mrn_val = row[mrn_col - 1] if mrn_col else None
            cms_val = row[cms_col - 1] if cms_col else None

            if name_val and str(name_val).strip():
                name_str = str(name_val).strip().lower()
                for name in placeholder_names:
                    if name in name_str:
                        row_issues.append(
                            {
                                "row": r,
                                "mrn": mrn_val,
                                "cms": cms_val,
                                "issue_type": "Possible Placeholder Name",
                                "description": f"Patient Name '{name_val}' may be a placeholder or test name",
                            }
                        )
                        break

    return issues, row_issues


# ---------------------------------------------------------------------------
# Email quality / suspicious-email detection
# ---------------------------------------------------------------------------

# Local-part prefixes that indicate opt-out, refusal, or placeholder emails
_SUSPICIOUS_LOCAL_PARTS = {
    "optout", "opt-out", "opt.out",
    "noreply", "no-reply", "no.reply",
    "donotcontact", "donotsend", "donotmail", "donotreply",
    "unsubscribe", "remove",
    "declined", "refused", "refuse",
    "none", "na", "n/a", "null", "void",
    "test", "testing", "tester",
    "fake", "bogus", "junk", "spam", "trash",
    "sample", "example", "demo", "placeholder",
    "noemail", "no-email", "no.email", "noemailaddress",
    "unknown", "notprovided", "notavailable",
    "abc", "asdf", "qwerty", "xxx", "zzz",
    "admin", "info", "noreply", "postmaster", "mailer-daemon",
}

# Well-known disposable / throwaway email domains (most common ones)
_DISPOSABLE_DOMAINS = {
    "mailinator.com", "guerrillamail.com", "tempmail.com", "throwaway.email",
    "yopmail.com", "sharklasers.com", "guerrillamailblock.com", "grr.la",
    "discard.email", "mailnesia.com", "maildrop.cc", "trashmail.com",
    "trashmail.me", "trashmail.net", "10minutemail.com", "temp-mail.org",
    "fakeinbox.com", "tempail.com", "tempr.email", "dispostable.com",
    "getnada.com", "emailondeck.com", "33mail.com", "mytemp.email",
    "mohmal.com", "burnermail.io", "inboxkitten.com",
}


def validate_email_quality(email_str):
    """Check an email address for suspicious / low-quality patterns.

    Returns a list of warning strings.  An empty list means the email looks OK.
    This is intentionally separate from *format* validation - it catches emails
    that are syntactically valid but semantically bogus.
    """
    warnings = []
    if not email_str:
        return warnings

    email_lower = email_str.strip().lower()
    local_part = email_lower.split("@")[0] if "@" in email_lower else email_lower
    domain = email_lower.split("@")[1] if "@" in email_lower else ""

    # 1. Exact-match or prefix match against suspicious local-part list
    if local_part in _SUSPICIOUS_LOCAL_PARTS:
        warnings.append(f"Potentially invalid local part '{local_part}' (may be opt-out / placeholder)")
    else:
        # Also check if local part *starts with* a suspicious prefix followed
        # by digits or punctuation, e.g. "optout1@", "test123@"
        for prefix in _SUSPICIOUS_LOCAL_PARTS:
            if local_part.startswith(prefix) and len(local_part) > len(prefix):
                remainder = local_part[len(prefix):]
                if all(c in "0123456789._-+" for c in remainder):
                    warnings.append(f"Potentially invalid local part '{local_part}' (resembles '{prefix}' + filler)")
                    break

    # 2. Single-character or all-numeric local part
    if len(local_part) == 1:
        warnings.append(f"Single-character local part '{local_part}'")
    elif local_part.isdigit():
        warnings.append(f"All-numeric local part '{local_part}'")

    # 3. Disposable / throwaway domain
    if domain in _DISPOSABLE_DOMAINS:
        warnings.append(f"Potentially disposable email domain '{domain}'")

    # 4. Very short overall address (e.g. "a@b.co" - 6 chars)
    if len(email_lower) <= 6 and not warnings:
        warnings.append(f"Potentially invalid - very short email address")

    return warnings


def check_email_quality_all_rows(sheet, email_col, mrn_col, cms_col):
    """Scan every row for suspicious email addresses.

    Returns two lists:
        cms1_issues  – list of dicts for CMS=1 rows (high priority)
        cms2_issues  – list of dicts for CMS=2 rows (informational)

    Each dict: {row, mrn, cms, email, warnings: [str, ...]}
    """
    cms1_issues = []
    cms2_issues = []

    if not email_col:
        return cms1_issues, cms2_issues

    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if is_blank_row(row):
            continue

        email_val = row[email_col - 1]
        if not email_val or not str(email_val).strip():
            continue

        email_str = str(email_val).strip()
        warnings = validate_email_quality(email_str)
        if not warnings:
            continue

        mrn_val = row[mrn_col - 1] if mrn_col else None
        cms_val = row[cms_col - 1] if cms_col else None

        entry = {
            "row": r,
            "mrn": mrn_val,
            "cms": cms_val,
            "email": email_str,
            "warnings": warnings,
        }

        cms_int = None
        try:
            if cms_val is not None and str(cms_val).strip():
                cms_int = int(float(str(cms_val).strip()))
        except (ValueError, TypeError):
            pass

        if cms_int == 2:
            cms2_issues.append(entry)
        else:
            cms1_issues.append(entry)

    return cms1_issues, cms2_issues


# ---------------------------------------------------------------------------
# People-search lookup helpers  (used by --lookup mode)
# ---------------------------------------------------------------------------

_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")


def collect_lookup_candidates(sheet, headers, mrn_col, cms_col):
    """
    Scan the OASCAPHS sheet for CMS=1 rows that need a manual people-search:
      - Invalid (non-blank) email address
      - No valid phone number (0 valid numbers across TELEPHONE + CELL PHONE)

    Rows with at least one valid phone but a bad one are flagged as "reference"
    (show the values, no search links). All others are "lookup" (show search links).

    Returns a list of dicts:
      {row, mrn, name, city, state, issues: [str, ...], mode, tel_value, cell_value}
    """
    email_col  = headers.get("EMAIL ADDRESS")
    tel_col    = headers.get("TELEPHONE")
    cell_col   = headers.get("CELL PHONE")
    name_col   = headers.get("PATIENT NAME")
    city_col   = headers.get("CITY")
    state_col  = headers.get("STATE")
    age_col    = headers.get("AGE")

    candidates = []

    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if is_blank_row(row):
            continue

        cms_val = row[cms_col - 1] if cms_col else None
        mrn_val = row[mrn_col - 1] if mrn_col else None

        # Only flag CMS=1 patients; use int(float(...)) to handle "1.0"-style cells
        try:
            cms_int = int(float(str(cms_val))) if cms_val is not None and str(cms_val).strip() else None
            if cms_int != 1:
                continue
        except (ValueError, TypeError):
            continue

        name_val  = str(row[name_col  - 1] or "").strip() if name_col  else ""
        city_val  = str(row[city_col  - 1] or "").strip() if city_col  else ""
        state_val = str(row[state_col - 1] or "").strip() if state_col else ""
        age_val   = row[age_col - 1] if age_col else None

        row_issues = []
        # "lookup"    → show people-search links (need to find contact info)
        # "reference" → show both phone values only (auditor corrects them manually)
        mode = "lookup"

        # Invalid (non-blank) email always triggers a lookup
        if email_col:
            email_val = row[email_col - 1]
            if email_val and str(email_val).strip():
                email_str = str(email_val).strip()
                if not _EMAIL_RE.match(email_str):
                    row_issues.append(f"Invalid email: {email_str}")

        # --- Phone logic ---
        tel_val  = row[tel_col  - 1] if tel_col  else None
        cell_val = row[cell_col - 1] if cell_col else None
        tel_str  = str(tel_val).strip()  if tel_val  else ""
        cell_str = str(cell_val).strip() if cell_val else ""

        tel_blank  = not tel_str
        cell_blank = not cell_str

        # Validate each present number
        def _phone_invalid(num_str):
            try:
                parsed = phonenumbers.parse(num_str, "US")
                return not phonenumbers.is_valid_number(parsed)
            except phonenumbers.NumberParseException:
                return True

        tel_invalid  = (not tel_blank)  and _phone_invalid(tel_str)
        cell_invalid = (not cell_blank) and _phone_invalid(cell_str)
        has_valid_phone = (not tel_blank and not tel_invalid) or (not cell_blank and not cell_invalid)

        phone_issues = []
        if not has_valid_phone:
            # 0 valid numbers - lookup needed
            if tel_blank and cell_blank:
                phone_issues.append("No phone number on file")
            else:
                if tel_blank:
                    phone_issues.append("No telephone on file")
                elif tel_invalid:
                    phone_issues.append(f"Invalid telephone: '{tel_str}'")
                if cell_blank:
                    phone_issues.append("No cell phone on file")
                elif cell_invalid:
                    phone_issues.append(f"Invalid cell phone: '{cell_str}'")
            # mode stays "lookup"
        else:
            # 1+ valid number - note any bad ones for reference, no lookup needed
            if tel_invalid:
                phone_issues.append(f"Invalid telephone: '{tel_str}' (cell phone is valid)")
            if cell_invalid:
                phone_issues.append(f"Invalid cell phone: '{cell_str}' (telephone is valid)")
            if phone_issues:
                mode = "reference"

        row_issues.extend(phone_issues)

        if row_issues:
            candidates.append({
                "row":       r,
                "mrn":       mrn_val,
                "name":      name_val,
                "age":       age_val,
                "city":      city_val,
                "state":     state_val,
                "issues":    row_issues,
                "mode":      mode,
                "tel_value": tel_str,
                "cell_value": cell_str,
            })

    return candidates


def build_person_search_urls(name: str, city: str = "", state: str = "") -> dict:
    """
    Return a dict of {site_label: url} with pre-populated people-search URLs.
    No network request is made here - links are lazy (only fetched on click).
    """
    from urllib.parse import quote, quote_plus

    name  = name.strip()
    city  = city.strip()
    state = state.strip()
    state_up = state.upper()

    name_q     = quote_plus(name)
    loc_q      = quote_plus(f"{city}, {state_up}".strip(", ")) if (city or state) else ""

    # URL-slug form: "John Smith" -> "john-smith"
    name_slug  = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    city_slug  = re.sub(r"[^a-z0-9]+", "-", city.lower()).strip("-")

    # WhitePages slug: "Barbara Arthur" -> "Barbara-Arthur", city+state: "Aurora-CO"
    wp_name_slug = re.sub(r"\s+", "-", name.strip())
    wp_city_state = f"{city}-{state_up}" if (city and state_up) else (city or state_up)
    wp_path = f"https://www.whitepages.com/name/{wp_name_slug}/{wp_city_state}" if wp_city_state else f"https://www.whitepages.com/name/{wp_name_slug}"
    wp_searched_name = quote(name.lower(), safe="")
    wp_searched_loc  = quote(f"{city}, {state_up}".strip(", "), safe="") if (city or state_up) else ""
    wp_query = f"fs=1&searchedName={wp_searched_name}"
    if wp_searched_loc:
        wp_query += f"&searchedLocation={wp_searched_loc}"

    urls = {}

    # WhitePages
    urls["WhitePages"] = f"{wp_path}?{wp_query}"

    # TruePeopleSearch
    tps_base = f"https://www.truepeoplesearch.com/results?name={name_q}"
    urls["TruePeopleSearch"] = f"{tps_base}&citystatezip={loc_q}" if loc_q else tps_base

    return urls
